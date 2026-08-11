from flask import (
    Flask, render_template, request, redirect, session, jsonify, send_file,
    send_from_directory
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
from urllib.parse import quote
from embarques import embarques_bp, crear_tablas as inicializar_embarques
from metricas import metricas_bp, crear_tablas as inicializar_metricas
from database import conectar, DatabaseError, IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from io import BytesIO
from datetime import datetime, time as hora_cero, timedelta
import json
import os
import threading
import time
import secrets
import hashlib
import hmac
import mimetypes
import re
import sqlite3

# Se indican los nombres reales de las carpetas. Windows no distingue mayúsculas,
# pero el servidor Linux de Render sí.
app = Flask(
    __name__,
    template_folder="Templates",
    static_folder="Static",
    static_url_path="/static",
)
app.register_blueprint(embarques_bp)
app.register_blueprint(metricas_bp)
app.secret_key = os.environ.get("SECRET_KEY", "clave-local-temporal-polo-oeste")
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = bool(os.environ.get("DATABASE_URL"))

USUARIO_ADMIN_INICIAL = os.environ.get("ADMIN_INITIAL_USER", "Emanuel Machado")
CONTRASENA_ADMIN_INICIAL = os.environ.get("ADMIN_INITIAL_PASSWORD", "123")

COLECTOR_PREDETERMINADO = "Handheld web"
INACTIVIDAD_MAXIMA = 30 * 60
RUTA_BASE = os.path.dirname(os.path.abspath(__file__))
# Permite ejecutar pruebas aisladas sin tocar la base operativa.
RUTA_BD = os.environ.get("POLO_OESTE_DB", os.path.join(RUTA_BASE, "actividad.db"))
RUTA_CERTIFICADOS = os.path.join(RUTA_BASE, "uploads", "certificados")
RUTA_ADJUNTOS_TAREAS = os.path.join(RUTA_BASE, "uploads", "tareas")


def almacenamiento_cloud():
    """En cloud los adjuntos van a PostgreSQL; el disco del servicio es temporal."""
    return bool(os.environ.get("DATABASE_URL"))


def conectar_db():
    return conectar(RUTA_BD)


def inicializar_db():
    conexion = conectar_db()

    try:
        conexion.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                usuario TEXT NOT NULL UNIQUE,
                contrasena_hash TEXT NOT NULL,
                es_admin INTEGER NOT NULL DEFAULT 0,
                nombre_funcionario TEXT,
                activo INTEGER NOT NULL DEFAULT 1
            )
        """)

        columnas_usuarios = {
            columna["name"]
            for columna in conexion.execute("PRAGMA table_info(usuarios)")
        }
        if "nombre_funcionario" not in columnas_usuarios:
            conexion.execute("ALTER TABLE usuarios ADD COLUMN nombre_funcionario TEXT")
        if "activo" not in columnas_usuarios:
            conexion.execute("ALTER TABLE usuarios ADD COLUMN activo INTEGER NOT NULL DEFAULT 1")

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS usuario_empresas (
                usuario_id INTEGER NOT NULL,
                empresa_codigo TEXT NOT NULL,
                origen TEXT NOT NULL DEFAULT 'Manual',
                asignado_en REAL NOT NULL,
                PRIMARY KEY (usuario_id, empresa_codigo),
                FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
            )
        """)

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS tareas_supervision (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supervisor_usuario TEXT NOT NULL,
                titulo TEXT NOT NULL,
                descripcion TEXT,
                fecha TEXT NOT NULL,
                hora TEXT,
                prioridad TEXT NOT NULL DEFAULT 'Media',
                estado TEXT NOT NULL DEFAULT 'Pendiente',
                creado_en REAL NOT NULL,
                actualizado_en REAL NOT NULL
            )
        """)
        conexion.execute("""
            CREATE INDEX IF NOT EXISTS indice_tareas_supervisor_fecha
            ON tareas_supervision(supervisor_usuario, fecha)
        """)

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS archivos_tareas_supervision (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tarea_id INTEGER NOT NULL,
                nombre_original TEXT NOT NULL,
                ruta_archivo TEXT NOT NULL,
                tamano INTEGER NOT NULL DEFAULT 0,
                subido_en REAL NOT NULL,
                FOREIGN KEY (tarea_id) REFERENCES tareas_supervision(id)
            )
        """)
        columnas_archivos = {
            columna["name"]
            for columna in conexion.execute("PRAGMA table_info(archivos_tareas_supervision)")
        }
        if "contenido" not in columnas_archivos:
            conexion.execute("ALTER TABLE archivos_tareas_supervision ADD COLUMN contenido BLOB")
        if "tipo_mime" not in columnas_archivos:
            conexion.execute("ALTER TABLE archivos_tareas_supervision ADD COLUMN tipo_mime TEXT")

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS planificaciones_horarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supervisor_usuario TEXT NOT NULL,
                empresa_codigo TEXT NOT NULL,
                semana_inicio TEXT NOT NULL,
                creado_en REAL NOT NULL,
                actualizado_en REAL NOT NULL,
                UNIQUE(supervisor_usuario, empresa_codigo, semana_inicio)
            )
        """)
        conexion.execute("""
            CREATE TABLE IF NOT EXISTS turnos_personal (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                planificacion_id INTEGER NOT NULL,
                usuario_id INTEGER NOT NULL,
                fecha TEXT NOT NULL,
                horario TEXT,
                UNIQUE(planificacion_id, usuario_id, fecha),
                FOREIGN KEY (planificacion_id) REFERENCES planificaciones_horarios(id),
                FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
            )
        """)

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS ausencias_funcionarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                usuario_id INTEGER NOT NULL,
                fecha_desde TEXT NOT NULL,
                fecha_hasta TEXT NOT NULL,
                tipo TEXT NOT NULL,
                detalle TEXT,
                archivo_nombre TEXT,
                archivo_ruta TEXT,
                creado_por TEXT NOT NULL,
                creado_en REAL NOT NULL,
                FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
            )
        """)
        columnas_ausencias = {
            columna["name"]
            for columna in conexion.execute("PRAGMA table_info(ausencias_funcionarios)")
        }
        if "archivo_contenido" not in columnas_ausencias:
            conexion.execute("ALTER TABLE ausencias_funcionarios ADD COLUMN archivo_contenido BLOB")
        if "archivo_mime" not in columnas_ausencias:
            conexion.execute("ALTER TABLE ausencias_funcionarios ADD COLUMN archivo_mime TEXT")

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS colectores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero TEXT NOT NULL UNIQUE,
                descripcion TEXT NOT NULL,
                activo INTEGER NOT NULL DEFAULT 1,
                clave_instalacion TEXT
            )
        """)

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS empresas (
                codigo TEXT PRIMARY KEY,
                descripcion TEXT NOT NULL
            )
        """)
        empresas_iniciales = [
            ("6009", "POLAKOF"), ("6000", "DECATHLON"),
            ("6007", "AVERIAS DECATHLON"), ("6005", "FORUM"),
            ("6004", "AVERIAS DECATHLON"), ("6003", "UNILOG"),
            ("6002", "KIABI"), ("6001", "TELESHOPPING")
        ]
        conexion.executemany("""
            INSERT OR IGNORE INTO empresas (codigo, descripcion) VALUES (?, ?)
        """, empresas_iniciales)

        columnas_colectores = {
            columna["name"]
            for columna in conexion.execute("PRAGMA table_info(colectores)")
        }

        if "clave_instalacion" not in columnas_colectores:
            conexion.execute("""
                ALTER TABLE colectores
                ADD COLUMN clave_instalacion TEXT
            """)

        # La clave que genera la app nunca se conserva en texto plano.  Las
        # instalaciones creadas con la versión anterior se migran sin perder
        # su vínculo actual.
        columnas_colectores = {
            columna["name"] for columna in conexion.execute("PRAGMA table_info(colectores)")
        }
        nuevas_columnas_colector = {
            "clave_instalacion_hash": "TEXT",
            "dispositivo_id": "TEXT",
            "emparejado_en": "REAL",
            "ultimo_ping": "REAL",
            "ultimo_movimiento": "REAL",
            "bateria_porcentaje": "INTEGER",
            "cargando": "INTEGER NOT NULL DEFAULT 0",
            "estado_dispositivo": "TEXT NOT NULL DEFAULT 'Sin conexión'"
        }
        for columna, definicion in nuevas_columnas_colector.items():
            if columna not in columnas_colectores:
                conexion.execute(f"ALTER TABLE colectores ADD COLUMN {columna} {definicion}")

        claves_anteriores = conexion.execute("""
            SELECT id, clave_instalacion FROM colectores
            WHERE (clave_instalacion_hash IS NULL OR clave_instalacion_hash = '')
              AND clave_instalacion IS NOT NULL AND clave_instalacion != ''
        """).fetchall()
        for colector in claves_anteriores:
            conexion.execute("""
                UPDATE colectores
                SET clave_instalacion_hash = ?, clave_instalacion = NULL
                WHERE id = ?
            """, (hash_token(colector["clave_instalacion"]), colector["id"]))

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS sesiones_handheld (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                token_hash TEXT NOT NULL UNIQUE,
                funcionario TEXT NOT NULL,
                colector TEXT NOT NULL,
                creado_en REAL NOT NULL,
                ultimo_ping REAL NOT NULL,
                revocada INTEGER NOT NULL DEFAULT 0
            )
        """)

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS actividades (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                funcionario TEXT NOT NULL,
                colector TEXT NOT NULL,
                estado TEXT NOT NULL,
                inicio REAL NOT NULL,
                ultimo_inicio REAL,
                tiempo_acumulado REAL NOT NULL DEFAULT 0,
                activa INTEGER NOT NULL DEFAULT 1,
                ultimo_ping REAL NOT NULL,
                finalizada_en REAL,
                origen_externo TEXT
            )
        """)

        columnas_actividades = {
            columna["name"]
            for columna in conexion.execute("PRAGMA table_info(actividades)")
        }

        if "origen_externo" not in columnas_actividades:
            conexion.execute("""
                ALTER TABLE actividades
                ADD COLUMN origen_externo TEXT
            """)

        conexion.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS indice_actividad_origen_externo
            ON actividades(origen_externo)
            WHERE origen_externo IS NOT NULL
        """)

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS periodos_actividad (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                actividad_id INTEGER NOT NULL,
                inicio REAL NOT NULL,
                fin REAL,
                FOREIGN KEY (actividad_id) REFERENCES actividades(id)
            )
        """)

        # Si se actualiza con una actividad que ya estaba en curso, se inicia
        # su primer período desde el último momento activo conocido.
        actividades_activas = conexion.execute("""
            SELECT id, ultimo_inicio
            FROM actividades
            WHERE activa = 1
              AND ultimo_inicio IS NOT NULL
              AND NOT EXISTS (
                  SELECT 1
                  FROM periodos_actividad
                  WHERE periodos_actividad.actividad_id = actividades.id
                    AND periodos_actividad.fin IS NULL
              )
        """).fetchall()

        for actividad in actividades_activas:
            conexion.execute("""
                INSERT INTO periodos_actividad (actividad_id, inicio)
                VALUES (?, ?)
            """, (actividad["id"], actividad["ultimo_inicio"]))

        admin_nuevo = conexion.execute("""
            SELECT id FROM usuarios
            WHERE usuario = ?
        """, (USUARIO_ADMIN_INICIAL,)).fetchone()

        admin_anterior = conexion.execute("""
            SELECT id FROM usuarios
            WHERE usuario = '123'
        """).fetchone()

        if admin_nuevo is None and admin_anterior is not None:
            conexion.execute("""
                UPDATE usuarios
                SET usuario = ?,
                    contrasena_hash = ?,
                    es_admin = 1
                WHERE usuario = '123'
            """, (
                USUARIO_ADMIN_INICIAL,
                generate_password_hash(CONTRASENA_ADMIN_INICIAL)
            ))

        elif admin_nuevo is None:
            conexion.execute("""
                INSERT INTO usuarios (usuario, contrasena_hash, es_admin)
                VALUES (?, ?, 1)
            """, (
                USUARIO_ADMIN_INICIAL,
                generate_password_hash(CONTRASENA_ADMIN_INICIAL)
            ))

        conexion.commit()
    finally:
        conexion.close()


def obtener_usuario(usuario):
    conexion = conectar_db()

    try:
        return conexion.execute("""
            SELECT *
            FROM usuarios
            WHERE usuario = ?
        """, (usuario,)).fetchone()
    finally:
        conexion.close()


def credenciales_correctas(usuario, contrasena):
    registro = obtener_usuario(usuario)

    return (
        registro is not None
        and bool(registro["activo"])
        and check_password_hash(registro["contrasena_hash"], contrasena)
    )


def usuario_es_admin():
    usuario = session.get("usuario")

    if not usuario:
        return False

    registro = obtener_usuario(usuario)
    return registro is not None and bool(registro["es_admin"]) and bool(registro["activo"])


def admin_requerido(funcion):
    @wraps(funcion)
    def envoltura(*args, **kwargs):
        if "usuario" not in session:
            return redirect("/")

        if not usuario_es_admin():
            return redirect("/principal")

        return funcion(*args, **kwargs)

    return envoltura


def redirigir_usuarios(mensaje):
    return redirect("/usuarios?mensaje=" + quote(mensaje))


def hash_token(token):
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def obtener_sesion_android():
    autorizacion = request.headers.get("Authorization", "")

    if not autorizacion.startswith("Bearer "):
        return None

    token = autorizacion[7:].strip()

    if not token:
        return None

    conexion = conectar_db()

    try:
        sesion_android = conexion.execute("""
            SELECT sesiones_handheld.*
            FROM sesiones_handheld
            INNER JOIN colectores
                ON colectores.numero = sesiones_handheld.colector
            WHERE sesiones_handheld.token_hash = ?
              AND sesiones_handheld.revocada = 0
              AND colectores.activo = 1
        """, (hash_token(token),)).fetchone()

        if sesion_android is not None:
            with conexion:
                conexion.execute("""
                    UPDATE sesiones_handheld
                    SET ultimo_ping = ?
                    WHERE id = ?
                """, (time.time(), sesion_android["id"]))

        return sesion_android
    finally:
        conexion.close()


def obtener_registro_abierto(conexion, funcionario):
    return conexion.execute("""
        SELECT *
        FROM actividades
        WHERE funcionario = ?
          AND estado != 'Actividad finalizada'
        ORDER BY id DESC
        LIMIT 1
    """, (funcionario,)).fetchone()


def calcular_tiempo(registro, ahora=None):
    if ahora is None:
        ahora = time.time()

    tiempo_total = float(registro["tiempo_acumulado"])

    if int(registro["activa"]) == 1 and registro["ultimo_inicio"] is not None:
        tiempo_total += max(0, ahora - float(registro["ultimo_inicio"]))

    return int(tiempo_total)


def formatear_tiempo(segundos):
    horas = segundos // 3600
    minutos = (segundos % 3600) // 60
    segundos_restantes = segundos % 60

    return f"{horas:02d}:{minutos:02d}:{segundos_restantes:02d}"


def fecha_desde_marca(marca_tiempo):
    return datetime.fromtimestamp(float(marca_tiempo)).strftime("%Y-%m-%d")


def limites_de_fecha(fecha_texto):
    try:
        fecha = datetime.strptime(fecha_texto, "%Y-%m-%d").date()
    except ValueError:
        return None

    inicio = datetime.combine(fecha, hora_cero.min).timestamp()
    fin = datetime.combine(fecha + timedelta(days=1), hora_cero.min).timestamp()

    return inicio, fin


def obtener_registros_actividad_filtrados(
    conexion, fecha="", fecha_desde="", fecha_hasta="",
    funcionario="", consulta="", ahora=None
):
    """Aplica filtros independientes o combinados al historial de actividad."""
    ahora = ahora or time.time()
    condiciones = []
    parametros = []

    limite_fecha = limites_de_fecha(fecha) if fecha else None
    limite_desde = limites_de_fecha(fecha_desde) if fecha_desde else None
    limite_hasta = limites_de_fecha(fecha_hasta) if fecha_hasta else None

    if limite_fecha:
        condiciones.append("inicio < ? AND COALESCE(finalizada_en, ?) >= ?")
        parametros.extend([limite_fecha[1], ahora, limite_fecha[0]])
    if limite_desde:
        condiciones.append("COALESCE(finalizada_en, ?) >= ?")
        parametros.extend([ahora, limite_desde[0]])
    if limite_hasta:
        condiciones.append("inicio < ?")
        parametros.append(limite_hasta[1])
    if funcionario:
        condiciones.append("funcionario LIKE ? COLLATE NOCASE")
        parametros.append("%" + funcionario + "%")
    if consulta:
        condiciones.append("""(
            funcionario LIKE ? COLLATE NOCASE OR
            colector LIKE ? COLLATE NOCASE OR
            estado LIKE ? COLLATE NOCASE OR
            CAST(id AS TEXT) LIKE ?
        )""")
        termino = "%" + consulta + "%"
        parametros.extend([termino, termino, termino, termino])

    where = " WHERE " + " AND ".join(condiciones) if condiciones else ""
    return conexion.execute(f"""
        SELECT * FROM actividades {where}
        ORDER BY
            CASE WHEN estado = 'Actividad finalizada' THEN 1 ELSE 0 END,
            id DESC
    """, parametros).fetchall()


def cerrar_periodo_activo(conexion, actividad_id, momento_final):
    conexion.execute("""
        UPDATE periodos_actividad
        SET fin = ?
        WHERE id = (
            SELECT id
            FROM periodos_actividad
            WHERE actividad_id = ?
              AND fin IS NULL
            ORDER BY id DESC
            LIMIT 1
        )
    """, (momento_final, actividad_id))


def finalizar_registro(conexion, registro, momento_final):
    tiempo_acumulado = float(registro["tiempo_acumulado"])

    if int(registro["activa"]) == 1 and registro["ultimo_inicio"] is not None:
        tiempo_acumulado += max(
            0,
            momento_final - float(registro["ultimo_inicio"])
        )

        cerrar_periodo_activo(conexion, registro["id"], momento_final)

    conexion.execute("""
        UPDATE actividades
        SET estado = 'Actividad finalizada',
            activa = 0,
            ultimo_inicio = NULL,
            tiempo_acumulado = ?,
            finalizada_en = ?,
            ultimo_ping = ?
        WHERE id = ?
    """, (
        tiempo_acumulado,
        momento_final,
        momento_final,
        registro["id"]
    ))


def cerrar_registros_inactivos(conexion, ahora):
    limite = ahora - INACTIVIDAD_MAXIMA

    registros = conexion.execute("""
        SELECT *
        FROM actividades
        WHERE estado != 'Actividad finalizada'
          AND ultimo_ping <= ?
    """, (limite,)).fetchall()

    for registro in registros:
        momento_final = float(registro["ultimo_ping"]) + INACTIVIDAD_MAXIMA
        finalizar_registro(conexion, registro, momento_final)


def estado_visible(registro):
    if registro["estado"] == "Actividad finalizada":
        return "Actividad finalizada"

    if int(registro["activa"]) == 1:
        return "En operacion (Activo)"

    return "En operacion (Pausado)"


def serializar_registro(registro, ahora=None):
    segundos = calcular_tiempo(registro, ahora)

    return {
        "id": registro["id"],
        "fecha": fecha_desde_marca(registro["inicio"]),
        "funcionario": registro["funcionario"],
        "status": estado_visible(registro),
        "tiempo": formatear_tiempo(segundos),
        "tiempo_segundos": segundos,
        "colector": registro["colector"],
        "activa": bool(registro["activa"])
    }


def finalizar_actividad_por_sesion(funcionario):
    conexion = conectar_db()

    try:
        ahora = time.time()

        with conexion:
            cerrar_registros_inactivos(conexion, ahora)
            registro = obtener_registro_abierto(conexion, funcionario)

            if registro is not None:
                finalizar_registro(conexion, registro, ahora)
    finally:
        conexion.close()


def monitor_de_inactividad():
    while True:
        time.sleep(60)
        conexion = None

        try:
            conexion = conectar_db()

            with conexion:
                cerrar_registros_inactivos(conexion, time.time())
        except (DatabaseError, sqlite3.Error):
            pass
        finally:
            if conexion is not None:
                conexion.close()


monitor_iniciado = False


def iniciar_monitor_de_inactividad():
    """Inicia un único monitor tanto local como en el servicio cloud."""
    global monitor_iniciado
    if monitor_iniciado:
        return
    hilo = threading.Thread(target=monitor_de_inactividad, daemon=True)
    hilo.start()
    monitor_iniciado = True


inicializar_db()
inicializar_embarques()
inicializar_metricas()
if os.environ.get("ENABLE_ACTIVITY_MONITOR") == "1":
    iniciar_monitor_de_inactividad()


@app.route("/health")
def health():
    return jsonify({"ok": True, "servicio": "Polo Oeste"})


@app.route("/service-worker.js")
def service_worker():
    respuesta = send_from_directory(RUTA_BASE, "Static/service-worker.js", mimetype="application/javascript")
    respuesta.headers["Cache-Control"] = "no-cache"
    return respuesta


def empresa_actual():
    return str(session.get("empresa_codigo", "")).strip()


@app.context_processor
def contexto_empresas():
    conexion = conectar_db()
    try:
        empresas = conexion.execute("""
            SELECT codigo, descripcion FROM empresas ORDER BY codigo ASC
        """).fetchall()
        return {
            "empresas_menu": empresas,
            "empresa_seleccionada": empresa_actual()
        }
    finally:
        conexion.close()


@app.route("/empresa/seleccionar", methods=["POST"])
def seleccionar_empresa():
    if "usuario" not in session:
        return redirect("/")
    codigo = request.form.get("empresa_codigo", "").strip()
    conexion = conectar_db()
    try:
        existe = not codigo or conexion.execute(
            "SELECT 1 FROM empresas WHERE codigo = ?", (codigo,)
        ).fetchone()
        if existe:
            session["empresa_codigo"] = codigo
    finally:
        conexion.close()
    return redirect(request.form.get("volver", "/principal"))


@app.route("/", methods=["GET", "POST"])
def inicio():
    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        contrasena = request.form.get("contrasena", "")

        if credenciales_correctas(usuario, contrasena):
            registro = obtener_usuario(usuario)
            if registro is not None and bool(registro["es_admin"]):
                session["usuario"] = usuario
                return redirect("/principal")

            # Las cuentas operativas no ingresan al panel de supervisiÃ³n:
            # se autentican directamente para utilizar el handheld web.
            session["handheld_usuario"] = usuario
            return redirect("/handheld/panel")

        return redirect("/?login=error")

    if "usuario" in session:
        return redirect("/principal")

    return render_template("Index.html")


@app.route("/logout", methods=["POST"])
def cerrar_sesion():
    session.pop("usuario", None)
    session.pop("handheld_usuario", None)
    return redirect("/")


@app.route("/principal")
@admin_requerido
def principal():
    return render_template(
        "principal.html",
        usuario=session.get("usuario")
    )


@app.route("/actividad")
@admin_requerido
def actividad():
    return render_template("actividad.html")


@app.route("/actividad/<int:actividad_id>/eliminar", methods=["POST"])
@admin_requerido
def eliminar_registro_actividad(actividad_id):
    conexion = conectar_db()

    try:
        with conexion:
            conexion.execute("""
                DELETE FROM periodos_actividad
                WHERE actividad_id = ?
            """, (actividad_id,))

            resultado = conexion.execute("""
                DELETE FROM actividades
                WHERE id = ?
            """, (actividad_id,))

        if resultado.rowcount == 0:
            return redirect("/actividad?mensaje=" + quote(
                "No se encontró el registro seleccionado."
            ))

        return redirect("/actividad?mensaje=" + quote(
            "Registro de actividad eliminado correctamente."
        ))
    finally:
        conexion.close()


def marca_de_exportacion(valor):
    """Convierte marcas Android (milisegundos) a segundos Unix."""
    marca = float(valor)
    return marca / 1000 if marca > 100_000_000_000 else marca


@app.route("/actividad/importar-handheld", methods=["POST"])
def importar_actividad_handheld():
    if "usuario" not in session:
        return redirect("/")

    archivo = request.files.get("archivo")

    if archivo is None or not archivo.filename:
        return redirect("/actividad?mensaje=" + quote(
            "Selecciona el archivo de actividad recibido por Bluetooth."
        ))

    try:
        datos = json.load(archivo.stream)
        sesiones = datos.get("sesiones", [])
    except (json.JSONDecodeError, UnicodeDecodeError):
        return redirect("/actividad?mensaje=" + quote(
            "El archivo seleccionado no es una exportación válida de Polo Oeste."
        ))

    if datos.get("version") != 1 or not isinstance(sesiones, list):
        return redirect("/actividad?mensaje=" + quote(
            "El formato del archivo no es compatible."
        ))

    conexion = conectar_db()
    recibidas = 0

    try:
        with conexion:
            for sesion_local in sesiones:
                try:
                    identificador = str(sesion_local["sesion_id"]).strip()
                    funcionario = str(sesion_local["funcionario"]).strip()
                    colector = str(sesion_local["colector"]).strip()
                    inicio = marca_de_exportacion(sesion_local["inicio"])
                    exportado_en = marca_de_exportacion(
                        sesion_local.get("exportado_en", time.time() * 1000)
                    )
                    periodos_locales = sesion_local["periodos"]
                except (KeyError, TypeError, ValueError):
                    continue

                if (
                    not identificador or len(identificador) > 120
                    or not funcionario or len(funcionario) > 100
                    or not colector or len(colector) > 100
                    or not isinstance(periodos_locales, list)
                ):
                    continue

                periodos = []
                inicio_abierto = None

                for periodo_local in periodos_locales:
                    try:
                        inicio_periodo = marca_de_exportacion(periodo_local["inicio"])
                        fin_local = periodo_local.get("fin")
                        fin_periodo = (
                            marca_de_exportacion(fin_local)
                            if fin_local is not None else None
                        )
                    except (KeyError, TypeError, ValueError):
                        continue

                    if inicio_periodo < inicio or (
                        fin_periodo is not None and fin_periodo < inicio_periodo
                    ):
                        continue

                    if fin_periodo is None:
                        inicio_abierto = inicio_periodo

                    periodos.append((inicio_periodo, fin_periodo))

                if not periodos:
                    continue

                activa = bool(sesion_local.get("activa", False))
                finalizada_local = sesion_local.get("finalizada_en")
                finalizada_en = (
                    marca_de_exportacion(finalizada_local)
                    if finalizada_local is not None else None
                )

                if not activa:
                    inicio_abierto = None

                tiempo_acumulado = sum(
                    max(0, fin - inicio_periodo)
                    for inicio_periodo, fin in periodos
                    if fin is not None
                )

                origen = "bluetooth:" + identificador
                existente = conexion.execute("""
                    SELECT id FROM actividades
                    WHERE origen_externo = ?
                """, (origen,)).fetchone()

                estado = (
                    "Actividad finalizada"
                    if finalizada_en is not None else "En operacion"
                )

                if existente is None:
                    cursor = conexion.execute("""
                        INSERT INTO actividades (
                            funcionario, colector, estado, inicio, ultimo_inicio,
                            tiempo_acumulado, activa, ultimo_ping, finalizada_en,
                            origen_externo
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        funcionario, colector, estado, inicio, inicio_abierto,
                        tiempo_acumulado, 1 if activa else 0, exportado_en,
                        finalizada_en, origen
                    ))
                    actividad_id = cursor.lastrowid
                else:
                    actividad_id = existente["id"]
                    conexion.execute("""
                        UPDATE actividades
                        SET funcionario = ?, colector = ?, estado = ?, inicio = ?,
                            ultimo_inicio = ?, tiempo_acumulado = ?, activa = ?,
                            ultimo_ping = ?, finalizada_en = ?
                        WHERE id = ?
                    """, (
                        funcionario, colector, estado, inicio, inicio_abierto,
                        tiempo_acumulado, 1 if activa else 0, exportado_en,
                        finalizada_en, actividad_id
                    ))
                    conexion.execute("""
                        DELETE FROM periodos_actividad
                        WHERE actividad_id = ?
                    """, (actividad_id,))

                for inicio_periodo, fin_periodo in periodos:
                    conexion.execute("""
                        INSERT INTO periodos_actividad (actividad_id, inicio, fin)
                        VALUES (?, ?, ?)
                    """, (actividad_id, inicio_periodo, fin_periodo))

                recibidas += 1

        if recibidas == 0:
            return redirect("/actividad?mensaje=" + quote(
                "No se encontraron períodos válidos en el archivo."
            ))

        return redirect("/actividad?mensaje=" + quote(
            f"Se importaron {recibidas} actividades desde el handheld."
        ))
    finally:
        conexion.close()


@app.route("/usuarios")
@admin_requerido
def usuarios():
    conexion = conectar_db()

    try:
        lista_usuarios = conexion.execute("""
            SELECT usuarios.id, usuarios.usuario, usuarios.es_admin, usuarios.activo,
                   usuarios.nombre_funcionario,
                   GROUP_CONCAT(usuario_empresas.empresa_codigo, '|') AS codigos_empresa
            FROM usuarios
            LEFT JOIN usuario_empresas ON usuario_empresas.usuario_id = usuarios.id
            GROUP BY usuarios.id
            ORDER BY usuarios.es_admin DESC, usuarios.usuario ASC
        """).fetchall()

        lista_empresas = conexion.execute("""
            SELECT codigo, descripcion FROM empresas ORDER BY codigo
        """).fetchall()

        lista_colectores = conexion.execute("""
            SELECT id, numero, descripcion, activo, dispositivo_id, emparejado_en,
                   ultimo_ping, ultimo_movimiento, bateria_porcentaje, cargando,
                   estado_dispositivo,
                   CASE WHEN clave_instalacion_hash IS NULL OR clave_instalacion_hash = ''
                        THEN 0 ELSE 1 END AS tiene_clave
            FROM colectores
            ORDER BY numero ASC
        """).fetchall()

        return render_template(
            "usuarios.html",
            usuarios=lista_usuarios,
            empresas=lista_empresas,
            colectores=lista_colectores,
            usuario_actual=session.get("usuario")
        )
    finally:
        conexion.close()


@app.route("/configuracion")
@admin_requerido
def configuracion():
    return redirect("/usuarios")


@app.route("/configuracion/apariencia")
@admin_requerido
def configuracion_apariencia():
    return render_template("apariencia.html")


@app.route("/configuracion/empresas")
@admin_requerido
def configuracion_empresas():
    conexion = conectar_db()
    try:
        empresas = conexion.execute("""
            SELECT empresas.codigo, empresas.descripcion,
                   COUNT(movimientos_metricas.id) AS registros
            FROM empresas
            LEFT JOIN movimientos_metricas
                ON movimientos_metricas.empresa_codigo = empresas.codigo
            GROUP BY empresas.codigo, empresas.descripcion
            ORDER BY empresas.codigo ASC
        """).fetchall()
        return render_template("empresas.html", empresas=empresas)
    finally:
        conexion.close()


@app.route("/configuracion/empresas/crear", methods=["POST"])
@admin_requerido
def crear_empresa():
    codigo = request.form.get("codigo", "").strip()
    descripcion = request.form.get("descripcion", "").strip().upper()
    if not codigo.isdigit() or not descripcion:
        return redirect("/configuracion/empresas?mensaje=" + quote("Ingresa un código numérico y una descripción."))
    conexion = conectar_db()
    try:
        with conexion:
            conexion.execute("INSERT INTO empresas (codigo, descripcion) VALUES (?, ?)", (codigo, descripcion))
        return redirect("/configuracion/empresas?mensaje=" + quote("Empresa creada correctamente."))
    except (IntegrityError, sqlite3.IntegrityError):
        return redirect("/configuracion/empresas?mensaje=" + quote("Ese código de empresa ya existe."))
    finally:
        conexion.close()


@app.route("/configuracion/empresas/<codigo>/eliminar", methods=["POST"])
@admin_requerido
def eliminar_empresa(codigo):
    conexion = conectar_db()
    try:
        tiene_datos = conexion.execute("""
            SELECT 1 FROM movimientos_metricas WHERE empresa_codigo = ? LIMIT 1
        """, (codigo,)).fetchone()
        if tiene_datos:
            return redirect("/configuracion/empresas?mensaje=" + quote(
                "No se puede eliminar una empresa que ya tiene datos cargados."
            ))
        with conexion:
            conexion.execute("DELETE FROM empresas WHERE codigo = ?", (codigo,))
        if session.get("empresa_codigo") == codigo:
            session["empresa_codigo"] = ""
        return redirect("/configuracion/empresas?mensaje=" + quote("Empresa eliminada."))
    finally:
        conexion.close()


def redirigir_organizador(vista, mensaje=""):
    destino = "/organizador?vista=" + vista
    if mensaje:
        destino += "&mensaje=" + quote(mensaje)
    return redirect(destino)


def empresas_de_funcionario(conexion, usuario_id):
    return conexion.execute("""
        SELECT empresas.codigo, empresas.descripcion, usuario_empresas.origen
        FROM usuario_empresas
        INNER JOIN empresas ON empresas.codigo = usuario_empresas.empresa_codigo
        WHERE usuario_empresas.usuario_id = ?
        ORDER BY empresas.codigo
    """, (usuario_id,)).fetchall()


def guardar_adjuntos_tarea(conexion, tarea_id, archivos):
    extensiones_permitidas = {".pdf", ".xlsx", ".xls", ".csv", ".docx", ".txt", ".jpg", ".jpeg", ".png", ".webp", ".zip"}
    guardados = 0
    for archivo in archivos:
        if archivo is None or not archivo.filename:
            continue
        nombre_seguro = secure_filename(archivo.filename)
        extension = os.path.splitext(nombre_seguro)[1].lower()
        if not nombre_seguro or extension not in extensiones_permitidas:
            continue
        nombre_interno = secrets.token_hex(16) + extension
        contenido = None
        tipo_mime = archivo.mimetype or mimetypes.guess_type(nombre_seguro)[0]
        if almacenamiento_cloud():
            contenido = archivo.read()
            tamanio = len(contenido)
        else:
            os.makedirs(RUTA_ADJUNTOS_TAREAS, exist_ok=True)
            destino = os.path.join(RUTA_ADJUNTOS_TAREAS, nombre_interno)
            archivo.save(destino)
            tamanio = os.path.getsize(destino)
        conexion.execute("""
            INSERT INTO archivos_tareas_supervision
            (tarea_id, nombre_original, ruta_archivo, tamano, subido_en, contenido, tipo_mime)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            tarea_id, nombre_seguro, nombre_interno, tamanio, time.time(),
            contenido, tipo_mime
        ))
        guardados += 1
    return guardados


@app.route("/organizador")
@admin_requerido
def organizador():
    vista = request.args.get("vista", "agenda").strip().lower()
    if vista not in {"agenda", "funcionarios", "ausencias"}:
        vista = "agenda"

    conexion = conectar_db()
    try:
        empresas = conexion.execute("""
            SELECT codigo, descripcion FROM empresas ORDER BY codigo
        """).fetchall()
        funcionarios = conexion.execute("""
            SELECT usuarios.id, usuarios.usuario,
                   COALESCE(NULLIF(usuarios.nombre_funcionario, ''), usuarios.usuario) AS nombre_funcionario,
                   usuarios.activo,
                   GROUP_CONCAT(empresas.codigo || ' · ' || empresas.descripcion, ' | ') AS operativas
            FROM usuarios
            LEFT JOIN usuario_empresas ON usuario_empresas.usuario_id = usuarios.id
            LEFT JOIN empresas ON empresas.codigo = usuario_empresas.empresa_codigo
            WHERE usuarios.es_admin = 0
            GROUP BY usuarios.id
            ORDER BY usuarios.activo DESC, nombre_funcionario COLLATE NOCASE ASC
        """).fetchall()
        asignaciones = {}
        for fila in conexion.execute("SELECT usuario_id, empresa_codigo FROM usuario_empresas"):
            asignaciones.setdefault(fila["usuario_id"], set()).add(fila["empresa_codigo"])
        ausencias = conexion.execute("""
            SELECT ausencias_funcionarios.*, usuarios.usuario,
                   COALESCE(NULLIF(usuarios.nombre_funcionario, ''), usuarios.usuario) AS funcionario
            FROM ausencias_funcionarios
            INNER JOIN usuarios ON usuarios.id = ausencias_funcionarios.usuario_id
            ORDER BY ausencias_funcionarios.fecha_desde DESC, ausencias_funcionarios.id DESC
            LIMIT 250
        """).fetchall()
        return render_template(
            "organizador.html", vista=vista, empresas=empresas,
            funcionarios=funcionarios, ausencias=ausencias,
            asignaciones=asignaciones,
            supervisor=session.get("usuario")
        )
    finally:
        conexion.close()


@app.route("/api/organizador/tareas")
@admin_requerido
def api_tareas_organizador():
    mes = request.args.get("mes", "").strip()
    if len(mes) != 7:
        mes = datetime.now().strftime("%Y-%m")
    conexion = conectar_db()
    try:
        tareas = conexion.execute("""
            SELECT id, titulo, descripcion, fecha, hora, prioridad, estado
            FROM tareas_supervision
            WHERE supervisor_usuario = ? AND substr(fecha, 1, 7) = ?
            ORDER BY fecha ASC, COALESCE(hora, '') ASC, id ASC
        """, (session.get("usuario"), mes)).fetchall()
        adjuntos = conexion.execute("""
            SELECT archivos_tareas_supervision.id, archivos_tareas_supervision.tarea_id,
                   archivos_tareas_supervision.nombre_original,
                   archivos_tareas_supervision.tamano
            FROM archivos_tareas_supervision
            INNER JOIN tareas_supervision
                ON tareas_supervision.id = archivos_tareas_supervision.tarea_id
            WHERE tareas_supervision.supervisor_usuario = ?
              AND substr(tareas_supervision.fecha, 1, 7) = ?
            ORDER BY archivos_tareas_supervision.id
        """, (session.get("usuario"), mes)).fetchall()
        adjuntos_por_tarea = {}
        for adjunto in adjuntos:
            datos = dict(adjunto)
            datos["url"] = f"/organizador/tareas/archivos/{adjunto['id']}/descargar"
            adjuntos_por_tarea.setdefault(adjunto["tarea_id"], []).append(datos)
        datos_tareas = []
        for tarea in tareas:
            dato = dict(tarea)
            dato["adjuntos"] = adjuntos_por_tarea.get(tarea["id"], [])
            datos_tareas.append(dato)
        return jsonify({"tareas": datos_tareas})
    finally:
        conexion.close()


@app.route("/organizador/tareas/guardar", methods=["POST"])
@admin_requerido
def guardar_tarea_organizador():
    identificador = request.form.get("id", "").strip()
    titulo = request.form.get("titulo", "").strip()
    descripcion = request.form.get("descripcion", "").strip()
    fecha = request.form.get("fecha", "").strip()
    hora = request.form.get("hora", "").strip()
    prioridad = request.form.get("prioridad", "Media").strip()
    estado = request.form.get("estado", "Pendiente").strip()
    if not titulo or len(titulo) > 120 or not fecha:
        return redirigir_organizador("agenda", "Completa el título y la fecha de la actividad.")
    if prioridad not in {"Baja", "Media", "Alta"}:
        prioridad = "Media"
    if estado not in {"Pendiente", "En curso", "Completada"}:
        estado = "Pendiente"
    ahora = time.time()
    archivos = request.files.getlist("adjuntos")
    conexion = conectar_db()
    try:
        with conexion:
            if identificador.isdigit():
                resultado = conexion.execute("""
                    UPDATE tareas_supervision
                    SET titulo = ?, descripcion = ?, fecha = ?, hora = ?,
                        prioridad = ?, estado = ?, actualizado_en = ?
                    WHERE id = ? AND supervisor_usuario = ?
                """, (titulo, descripcion, fecha, hora or None, prioridad, estado,
                      ahora, int(identificador), session.get("usuario")))
                if not resultado.rowcount:
                    return redirigir_organizador("agenda", "No se encontró la actividad.")
                tarea_id = int(identificador)
                mensaje = "Actividad actualizada."
            else:
                cursor = conexion.execute("""
                    INSERT INTO tareas_supervision (
                        supervisor_usuario, titulo, descripcion, fecha, hora,
                        prioridad, estado, creado_en, actualizado_en
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (session.get("usuario"), titulo, descripcion, fecha, hora or None,
                      prioridad, estado, ahora, ahora))
                tarea_id = cursor.lastrowid
                mensaje = "Actividad creada."
            adjuntos_guardados = guardar_adjuntos_tarea(conexion, tarea_id, archivos)
            if adjuntos_guardados:
                mensaje += f" Se adjuntaron {adjuntos_guardados} archivo(s)."
        return redirigir_organizador("agenda", mensaje)
    finally:
        conexion.close()


@app.route("/organizador/tareas/<int:tarea_id>/eliminar", methods=["POST"])
@admin_requerido
def eliminar_tarea_organizador(tarea_id):
    conexion = conectar_db()
    try:
        with conexion:
            conexion.execute("""
                DELETE FROM tareas_supervision
                WHERE id = ? AND supervisor_usuario = ?
            """, (tarea_id, session.get("usuario")))
        return redirigir_organizador("agenda", "Actividad eliminada.")
    finally:
        conexion.close()


@app.route("/organizador/tareas/archivos/<int:archivo_id>/descargar")
@admin_requerido
def descargar_adjunto_tarea(archivo_id):
    conexion = conectar_db()
    try:
        adjunto = conexion.execute("""
            SELECT archivos_tareas_supervision.nombre_original,
                   archivos_tareas_supervision.ruta_archivo,
                   archivos_tareas_supervision.contenido,
                   archivos_tareas_supervision.tipo_mime
            FROM archivos_tareas_supervision
            INNER JOIN tareas_supervision
                ON tareas_supervision.id = archivos_tareas_supervision.tarea_id
            WHERE archivos_tareas_supervision.id = ?
              AND tareas_supervision.supervisor_usuario = ?
        """, (archivo_id, session.get("usuario"))).fetchone()
        if adjunto is None:
            return redirect("/organizador")
        if adjunto["contenido"] is not None:
            return send_file(
                BytesIO(adjunto["contenido"]), as_attachment=True,
                download_name=adjunto["nombre_original"],
                mimetype=adjunto["tipo_mime"] or "application/octet-stream"
            )
        return send_from_directory(
            RUTA_ADJUNTOS_TAREAS, adjunto["ruta_archivo"],
            as_attachment=True, download_name=adjunto["nombre_original"]
        )
    finally:
        conexion.close()


def lunes_de_semana(fecha_texto=""):
    try:
        referencia = datetime.strptime(fecha_texto, "%Y-%m-%d").date()
    except ValueError:
        hoy = datetime.now().date()
        dias_hasta_lunes = 7 - hoy.weekday()
        referencia = hoy + timedelta(days=dias_hasta_lunes)
    return referencia - timedelta(days=referencia.weekday())


@app.route("/organizador/horarios")
@admin_requerido
def horarios_personal():
    semana_inicio = lunes_de_semana(request.args.get("semana", ""))
    semana_fin = semana_inicio + timedelta(days=6)
    dias = [semana_inicio + timedelta(days=indice) for indice in range(7)]
    conexion = conectar_db()
    try:
        empresas = conexion.execute("SELECT codigo, descripcion FROM empresas ORDER BY codigo").fetchall()
        codigo_empresa = request.args.get("empresa", "").strip()
        codigos_validos = {empresa["codigo"] for empresa in empresas}
        if codigo_empresa not in codigos_validos:
            codigo_empresa = empresa_actual() if empresa_actual() in codigos_validos else (empresas[0]["codigo"] if empresas else "")

        funcionarios = conexion.execute("""
            SELECT DISTINCT usuarios.id, usuarios.usuario,
                   COALESCE(NULLIF(usuarios.nombre_funcionario, ''), usuarios.usuario) AS nombre_funcionario
            FROM usuarios
            INNER JOIN usuario_empresas ON usuario_empresas.usuario_id = usuarios.id
            WHERE usuarios.es_admin = 0
              AND usuarios.activo = 1
              AND usuario_empresas.empresa_codigo = ?
            ORDER BY nombre_funcionario COLLATE NOCASE
        """, (codigo_empresa,)).fetchall()
        ausencias = conexion.execute("""
            SELECT usuario_id, fecha_desde, fecha_hasta, tipo, detalle
            FROM ausencias_funcionarios
            WHERE fecha_desde <= ? AND fecha_hasta >= ?
        """, (semana_fin.isoformat(), semana_inicio.isoformat())).fetchall()
        ausencias_por_dia = {}
        for ausencia in ausencias:
            inicio = max(datetime.strptime(ausencia["fecha_desde"], "%Y-%m-%d").date(), semana_inicio)
            fin = min(datetime.strptime(ausencia["fecha_hasta"], "%Y-%m-%d").date(), semana_fin)
            for indice in range((fin - inicio).days + 1):
                fecha = (inicio + timedelta(days=indice)).isoformat()
                texto_ausencia = ausencia["tipo"] + (": " + ausencia["detalle"] if ausencia["detalle"] else "")
                ausencias_por_dia[f"{ausencia['usuario_id']}:{fecha}"] = texto_ausencia

        planificacion = conexion.execute("""
            SELECT id FROM planificaciones_horarios
            WHERE supervisor_usuario = ? AND empresa_codigo = ? AND semana_inicio = ?
        """, (session.get("usuario"), codigo_empresa, semana_inicio.isoformat())).fetchone()
        turnos = {}
        if planificacion:
            for turno in conexion.execute("""
                SELECT usuario_id, fecha, horario FROM turnos_personal WHERE planificacion_id = ?
            """, (planificacion["id"],)):
                turnos[f"{turno['usuario_id']}:{turno['fecha']}"] = turno["horario"] or ""

        hitos = conexion.execute("""
            SELECT fecha, proveedor, barco, estado
            FROM embarques
            WHERE fecha >= ? AND fecha <= ?
              AND empresa_codigo = ?
            ORDER BY fecha, proveedor
        """, (semana_inicio.isoformat(), semana_fin.isoformat(), codigo_empresa)).fetchall()
        return render_template(
            "horarios.html", empresas=empresas, codigo_empresa=codigo_empresa,
            semana_inicio=semana_inicio.isoformat(), semana_fin=semana_fin.isoformat(),
            dias=dias, funcionarios=funcionarios, turnos=turnos,
            ausencias_por_dia=ausencias_por_dia, hitos=hitos,
            nombres_dias=["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
        )
    finally:
        conexion.close()


@app.route("/organizador/horarios/guardar", methods=["POST"])
@admin_requerido
def guardar_horarios_personal():
    codigo_empresa = request.form.get("empresa", "").strip()
    semana_inicio = lunes_de_semana(request.form.get("semana_inicio", ""))
    semana_fin = semana_inicio + timedelta(days=6)
    conexion = conectar_db()
    try:
        empresa = conexion.execute("SELECT codigo FROM empresas WHERE codigo = ?", (codigo_empresa,)).fetchone()
        if empresa is None:
            return redirect("/organizador/horarios")
        funcionarios = conexion.execute("""
            SELECT DISTINCT usuarios.id
            FROM usuarios
            INNER JOIN usuario_empresas ON usuario_empresas.usuario_id = usuarios.id
            WHERE usuarios.es_admin = 0 AND usuarios.activo = 1
              AND usuario_empresas.empresa_codigo = ?
        """, (codigo_empresa,)).fetchall()
        ausencias = conexion.execute("""
            SELECT usuario_id, fecha_desde, fecha_hasta FROM ausencias_funcionarios
            WHERE fecha_desde <= ? AND fecha_hasta >= ?
        """, (semana_fin.isoformat(), semana_inicio.isoformat())).fetchall()
        ausentes = set()
        for ausencia in ausencias:
            inicio = max(datetime.strptime(ausencia["fecha_desde"], "%Y-%m-%d").date(), semana_inicio)
            fin = min(datetime.strptime(ausencia["fecha_hasta"], "%Y-%m-%d").date(), semana_fin)
            for indice in range((fin - inicio).days + 1):
                ausentes.add((ausencia["usuario_id"], (inicio + timedelta(days=indice)).isoformat()))
        ahora = time.time()
        with conexion:
            conexion.execute("""
                INSERT OR IGNORE INTO planificaciones_horarios
                (supervisor_usuario, empresa_codigo, semana_inicio, creado_en, actualizado_en)
                VALUES (?, ?, ?, ?, ?)
            """, (session.get("usuario"), codigo_empresa, semana_inicio.isoformat(), ahora, ahora))
            planificacion = conexion.execute("""
                SELECT id FROM planificaciones_horarios
                WHERE supervisor_usuario = ? AND empresa_codigo = ? AND semana_inicio = ?
            """, (session.get("usuario"), codigo_empresa, semana_inicio.isoformat())).fetchone()
            conexion.execute("UPDATE planificaciones_horarios SET actualizado_en = ? WHERE id = ?", (ahora, planificacion["id"]))
            for funcionario in funcionarios:
                for indice in range(7):
                    fecha = (semana_inicio + timedelta(days=indice)).isoformat()
                    if (funcionario["id"], fecha) in ausentes:
                        conexion.execute("""
                            DELETE FROM turnos_personal WHERE planificacion_id = ? AND usuario_id = ? AND fecha = ?
                        """, (planificacion["id"], funcionario["id"], fecha))
                        continue
                    valor = request.form.get(f"turno_{funcionario['id']}_{fecha}", "").strip()[:50]
                    conexion.execute("""
                        INSERT INTO turnos_personal (planificacion_id, usuario_id, fecha, horario)
                        VALUES (?, ?, ?, ?)
                        ON CONFLICT(planificacion_id, usuario_id, fecha)
                        DO UPDATE SET horario = excluded.horario
                    """, (planificacion["id"], funcionario["id"], fecha, valor))
        return redirect("/organizador/horarios?empresa=" + quote(codigo_empresa) + "&semana=" + semana_inicio.isoformat() + "&mensaje=" + quote("Horario guardado."))
    finally:
        conexion.close()


@app.route("/organizador/funcionarios/<int:usuario_id>/operativas", methods=["POST"])
@admin_requerido
def actualizar_operativas_funcionario(usuario_id):
    codigos = [codigo.strip() for codigo in request.form.getlist("empresas") if codigo.strip()]
    nombre_funcionario = request.form.get("nombre_funcionario", "").strip()
    activo = 1 if request.form.get("activo") == "1" else 0
    conexion = conectar_db()
    try:
        usuario = conexion.execute("""
            SELECT id FROM usuarios WHERE id = ? AND es_admin = 0
        """, (usuario_id,)).fetchone()
        existentes = {
            fila["codigo"] for fila in conexion.execute("SELECT codigo FROM empresas")
        }
        codigos = [codigo for codigo in codigos if codigo in existentes]
        if usuario is None:
            return redirigir_organizador("funcionarios", "Funcionario no encontrado.")
        with conexion:
            conexion.execute("""
                UPDATE usuarios SET nombre_funcionario = ?, activo = ? WHERE id = ?
            """, (nombre_funcionario or None, activo, usuario_id))
            conexion.execute("DELETE FROM usuario_empresas WHERE usuario_id = ? AND origen = 'Manual'", (usuario_id,))
            conexion.executemany("""
                INSERT OR IGNORE INTO usuario_empresas
                (usuario_id, empresa_codigo, origen, asignado_en)
                VALUES (?, ?, 'Manual', ?)
            """, [(usuario_id, codigo, time.time()) for codigo in codigos])
        return redirigir_organizador("funcionarios", "Ficha del funcionario actualizada.")
    finally:
        conexion.close()


@app.route("/organizador/ausencias/crear", methods=["POST"])
@admin_requerido
def crear_ausencia_funcionario():
    try:
        usuario_id = int(request.form.get("usuario_id", ""))
    except ValueError:
        usuario_id = 0
    fecha_desde = request.form.get("fecha_desde", "").strip()
    fecha_hasta = request.form.get("fecha_hasta", "").strip()
    tipo = request.form.get("tipo", "Falta").strip()
    detalle = request.form.get("detalle", "").strip()
    archivo = request.files.get("certificado")
    if not usuario_id or not fecha_desde or not fecha_hasta or fecha_hasta < fecha_desde:
        return redirigir_organizador("ausencias", "Revisa el funcionario y el período de la ausencia.")
    if tipo not in {"Falta", "Certificación médica", "Licencia", "Otro"}:
        tipo = "Otro"

    archivo_nombre = None
    archivo_ruta = None
    archivo_contenido = None
    archivo_mime = None
    if archivo is not None and archivo.filename:
        extension = os.path.splitext(secure_filename(archivo.filename))[1].lower()
        if extension not in {".jpg", ".jpeg", ".png", ".webp"}:
            return redirigir_organizador("ausencias", "El certificado debe ser una imagen JPG, PNG o WEBP.")
        archivo_nombre = secure_filename(archivo.filename)
        archivo_ruta = f"{secrets.token_hex(12)}{extension}"
        archivo_mime = archivo.mimetype or mimetypes.guess_type(archivo_nombre)[0]
        if almacenamiento_cloud():
            archivo_contenido = archivo.read()
        else:
            os.makedirs(RUTA_CERTIFICADOS, exist_ok=True)
            archivo.save(os.path.join(RUTA_CERTIFICADOS, archivo_ruta))

    conexion = conectar_db()
    try:
        usuario = conexion.execute("SELECT id FROM usuarios WHERE id = ? AND es_admin = 0", (usuario_id,)).fetchone()
        if usuario is None:
            return redirigir_organizador("ausencias", "El funcionario seleccionado no es válido.")
        with conexion:
            conexion.execute("""
                INSERT INTO ausencias_funcionarios
                (usuario_id, fecha_desde, fecha_hasta, tipo, detalle, archivo_nombre,
                 archivo_ruta, creado_por, creado_en, archivo_contenido, archivo_mime)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (usuario_id, fecha_desde, fecha_hasta, tipo, detalle, archivo_nombre,
                  archivo_ruta, session.get("usuario"), time.time(), archivo_contenido,
                  archivo_mime))
        return redirigir_organizador("ausencias", "Ausencia registrada correctamente.")
    finally:
        conexion.close()


@app.route("/organizador/ausencias/<int:ausencia_id>/certificado")
@admin_requerido
def ver_certificado_ausencia(ausencia_id):
    conexion = conectar_db()
    try:
        ausencia = conexion.execute("""
            SELECT archivo_ruta, archivo_nombre, archivo_contenido, archivo_mime
            FROM ausencias_funcionarios WHERE id = ?
        """, (ausencia_id,)).fetchone()
        if ausencia is None or not ausencia["archivo_ruta"]:
            return redirect("/organizador?vista=ausencias")
        if ausencia["archivo_contenido"] is not None:
            return send_file(
                BytesIO(ausencia["archivo_contenido"]),
                download_name=ausencia["archivo_nombre"] or "certificado",
                mimetype=ausencia["archivo_mime"] or "application/octet-stream"
            )
        return send_from_directory(RUTA_CERTIFICADOS, ausencia["archivo_ruta"])
    finally:
        conexion.close()


@app.route("/usuarios/exportar")
@admin_requerido
def exportar_usuarios():
    filtro_usuarios = request.args.get("usuarios", "").strip()
    filtro_colectores = request.args.get("colectores", "").strip()
    conexion = conectar_db()
    try:
        usuarios_db = conexion.execute("""
            SELECT usuario, es_admin FROM usuarios
            WHERE usuario LIKE ? COLLATE NOCASE
            ORDER BY es_admin DESC, usuario ASC
        """, ("%" + filtro_usuarios + "%",)).fetchall()
        colectores_db = conexion.execute("""
            SELECT numero, descripcion, activo FROM colectores
            WHERE numero LIKE ? COLLATE NOCASE OR descripcion LIKE ? COLLATE NOCASE
            ORDER BY numero ASC
        """, ("%" + filtro_colectores + "%", "%" + filtro_colectores + "%")).fetchall()
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Usuarios"
        hoja.append(["Usuario", "Tipo"])
        for registro in usuarios_db:
            hoja.append([registro["usuario"], "Administrador" if registro["es_admin"] else "Operario"])
        equipos = libro.create_sheet("Handhelds")
        equipos.append(["Número", "Descripción", "Estado"])
        for registro in colectores_db:
            equipos.append([registro["numero"], registro["descripcion"], "Activo" if registro["activo"] else "Inactivo"])
        for pagina in (hoja, equipos):
            pagina.freeze_panes = "A2"
            for celda in pagina[1]:
                celda.fill = PatternFill("solid", fgColor="315F36")
                celda.font = Font(color="FFFFFF", bold=True)
            for letra, ancho in zip("ABC", [25, 34, 16]):
                pagina.column_dimensions[letra].width = ancho
        archivo = BytesIO()
        libro.save(archivo)
        archivo.seek(0)
        return send_file(
            archivo, as_attachment=True, download_name="usuarios_y_handhelds.xlsx",
            mimetype=("application/vnd.openxmlformats-officedocument."
                      "spreadsheetml.sheet")
        )
    finally:
        conexion.close()


@app.route("/usuarios/crear", methods=["POST"])
@admin_requerido
def crear_usuario():
    usuario = request.form.get("usuario", "").strip()
    contrasena = request.form.get("contrasena", "")
    es_admin = 1 if request.form.get("tipo") == "admin" else 0
    nombre_funcionario = request.form.get("nombre_funcionario", "").strip()
    codigos_empresa = [codigo.strip() for codigo in request.form.getlist("empresas") if codigo.strip()]

    if len(usuario) < 3 or len(contrasena) < 3:
        return redirigir_usuarios(
            "El usuario y la contraseña deben tener al menos 3 caracteres."
        )

    conexion = conectar_db()

    try:
        with conexion:
            cursor = conexion.execute("""
                INSERT INTO usuarios (usuario, contrasena_hash, es_admin, nombre_funcionario)
                VALUES (?, ?, ?, ?)
            """, (
                usuario,
                generate_password_hash(contrasena),
                es_admin,
                nombre_funcionario or None
            ))
            existentes = {fila["codigo"] for fila in conexion.execute("SELECT codigo FROM empresas")}
            if not es_admin:
                conexion.executemany("""
                    INSERT OR IGNORE INTO usuario_empresas
                    (usuario_id, empresa_codigo, origen, asignado_en)
                    VALUES (?, ?, 'Manual', ?)
                """, [(cursor.lastrowid, codigo, time.time()) for codigo in codigos_empresa if codigo in existentes])

        return redirigir_usuarios("Usuario creado correctamente.")
    except (IntegrityError, sqlite3.IntegrityError):
        return redirigir_usuarios("Ese nombre de usuario ya existe.")
    finally:
        conexion.close()


@app.route("/usuarios/actualizar", methods=["POST"])
@admin_requerido
def actualizar_usuario():
    identificador = request.form.get("id", "")
    nuevo_usuario = request.form.get("usuario", "").strip()
    nueva_contrasena = request.form.get("contrasena", "")
    es_admin = 1 if request.form.get("tipo") == "admin" else 0
    nombre_funcionario = request.form.get("nombre_funcionario", "").strip()

    if len(nuevo_usuario) < 3:
        return redirigir_usuarios("El usuario debe tener al menos 3 caracteres.")

    conexion = conectar_db()

    try:
        usuario_objetivo = conexion.execute("""
            SELECT * FROM usuarios WHERE id = ?
        """, (identificador,)).fetchone()

        if usuario_objetivo is None:
            return redirigir_usuarios("No se encontró el usuario seleccionado.")

        if (
            usuario_objetivo["usuario"] == session.get("usuario")
            and es_admin == 0
        ):
            return redirigir_usuarios(
                "No puedes quitarte el permiso de administrador."
            )

        with conexion:
            if nueva_contrasena:
                conexion.execute("""
                    UPDATE usuarios
                    SET usuario = ?,
                        contrasena_hash = ?,
                        es_admin = ?,
                        nombre_funcionario = ?
                    WHERE id = ?
                """, (
                    nuevo_usuario,
                    generate_password_hash(nueva_contrasena),
                    es_admin,
                    nombre_funcionario or None,
                    identificador
                ))
            else:
                conexion.execute("""
                    UPDATE usuarios
                    SET usuario = ?,
                        es_admin = ?,
                        nombre_funcionario = ?
                    WHERE id = ?
                """, (
                    nuevo_usuario,
                    es_admin,
                    nombre_funcionario or None,
                    identificador
                ))

        if usuario_objetivo["usuario"] == session.get("usuario"):
            session["usuario"] = nuevo_usuario

        return redirigir_usuarios("Usuario actualizado correctamente.")
    except (IntegrityError, sqlite3.IntegrityError):
        return redirigir_usuarios("Ese nombre de usuario ya existe.")
    finally:
        conexion.close()


@app.route("/usuarios/eliminar", methods=["POST"])
@admin_requerido
def eliminar_usuario():
    identificador = request.form.get("id", "")
    conexion = conectar_db()

    try:
        usuario_objetivo = conexion.execute("""
            SELECT * FROM usuarios WHERE id = ?
        """, (identificador,)).fetchone()

        if usuario_objetivo is None:
            return redirigir_usuarios("No se encontró el usuario seleccionado.")

        if usuario_objetivo["usuario"] == session.get("usuario"):
            return redirigir_usuarios("No puedes eliminar tu propio usuario.")

        if int(usuario_objetivo["es_admin"]) == 1:
            cantidad_admins = conexion.execute("""
                SELECT COUNT(*) AS total
                FROM usuarios
                WHERE es_admin = 1
            """).fetchone()["total"]

            if cantidad_admins <= 1:
                return redirigir_usuarios(
                    "Debe existir al menos un administrador."
                )

        with conexion:
            conexion.execute("""
                DELETE FROM usuarios WHERE id = ?
            """, (identificador,))

        return redirigir_usuarios("Usuario eliminado correctamente.")
    finally:
        conexion.close()


@app.route("/colectores/crear", methods=["POST"])
@admin_requerido
def crear_colector():
    numero = request.form.get("numero", "").strip()
    descripcion = request.form.get("descripcion", "").strip()
    codigo_instalacion = request.form.get("codigo_instalacion", "").strip()

    if not numero or not descripcion or len(codigo_instalacion) < 12:
        return redirigir_usuarios(
            "El número y la descripción del colector son obligatorios."
        )

    conexion = conectar_db()

    try:
        with conexion:
            conexion.execute("""
                INSERT INTO colectores (
                    numero, descripcion, activo, clave_instalacion_hash,
                    estado_dispositivo
                )
                VALUES (?, ?, 1, ?, 'Pendiente de enlace')
            """, (numero, descripcion, hash_token(codigo_instalacion)))

        return redirigir_usuarios("Handheld creado correctamente.")
    except (IntegrityError, sqlite3.IntegrityError):
        return redirigir_usuarios("Ese número de colector ya existe.")
    finally:
        conexion.close()


@app.route("/colectores/actualizar", methods=["POST"])
@admin_requerido
def actualizar_colector():
    identificador = request.form.get("id", "")
    numero = request.form.get("numero", "").strip()
    descripcion = request.form.get("descripcion", "").strip()
    activo = 1 if request.form.get("activo") == "1" else 0

    if not numero or not descripcion:
        return redirigir_usuarios(
            "El número y la descripción del colector son obligatorios."
        )

    conexion = conectar_db()

    try:
        with conexion:
            conexion.execute("""
                UPDATE colectores
                SET numero = ?,
                    descripcion = ?,
                    activo = ?
                WHERE id = ?
            """, (
                numero,
                descripcion,
                activo,
                identificador
            ))

        return redirigir_usuarios("Handheld actualizado correctamente.")
    except (IntegrityError, sqlite3.IntegrityError):
        return redirigir_usuarios("Ese número de colector ya existe.")
    finally:
        conexion.close()


@app.route("/colectores/eliminar", methods=["POST"])
@admin_requerido
def eliminar_colector():
    identificador = request.form.get("id", "")
    conexion = conectar_db()

    try:
        with conexion:
            conexion.execute("""
                DELETE FROM colectores WHERE id = ?
            """, (identificador,))

        return redirigir_usuarios("Handheld eliminado correctamente.")
    finally:
        conexion.close()


@app.route("/handheld", methods=["GET", "POST"])
def handheld_login():
    if "handheld_usuario" in session:
        return redirect("/handheld/panel")

    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        contrasena = request.form.get("contrasena", "")

        registro = obtener_usuario(usuario)
        if credenciales_correctas(usuario, contrasena) and registro is not None and not bool(registro["es_admin"]):
            session["handheld_usuario"] = usuario
            return redirect("/handheld/panel")

        return redirect("/handheld?login=error")

    return render_template("handheld_login.html")


@app.route("/handheld/panel")
def handheld_panel():
    if "handheld_usuario" not in session:
        return redirect("/handheld")

    return render_template("handheld.html")


@app.route("/handheld/logout", methods=["POST"])
def handheld_logout():
    funcionario = session.get("handheld_usuario")

    if funcionario:
        finalizar_actividad_por_sesion(funcionario)

    session.pop("handheld_usuario", None)
    return redirect("/handheld?logout=ok")


@app.route("/api/handheld/actividad")
def api_handheld_actividad():
    funcionario = session.get("handheld_usuario")

    if not funcionario:
        return jsonify({"error": "Sesion no valida"}), 401

    conexion = conectar_db()

    try:
        ahora = time.time()

        with conexion:
            cerrar_registros_inactivos(conexion, ahora)
            registro = obtener_registro_abierto(conexion, funcionario)

            if registro is None:
                return jsonify({
                    "tiene_registro": False,
                    "activa": False,
                    "tiempo_segundos": 0,
                    "tiempo": "00:00:00"
                })

            conexion.execute("""
                UPDATE actividades
                SET ultimo_ping = ?
                WHERE id = ?
            """, (ahora, registro["id"]))

            registro = obtener_registro_abierto(conexion, funcionario)
            respuesta = serializar_registro(registro, ahora)
            respuesta["tiene_registro"] = True

            return jsonify(respuesta)
    finally:
        conexion.close()


@app.route("/api/handheld/actividad/toggle", methods=["POST"])
def api_toggle_actividad():
    funcionario = session.get("handheld_usuario")

    if not funcionario:
        return jsonify({"error": "Sesion no valida"}), 401

    conexion = conectar_db()

    try:
        ahora = time.time()

        with conexion:
            cerrar_registros_inactivos(conexion, ahora)
            registro = obtener_registro_abierto(conexion, funcionario)

            if registro is None:
                cursor = conexion.execute("""
                    INSERT INTO actividades (
                        funcionario,
                        colector,
                        estado,
                        inicio,
                        ultimo_inicio,
                        tiempo_acumulado,
                        activa,
                        ultimo_ping
                    )
                    VALUES (?, ?, 'En operacion', ?, ?, 0, 1, ?)
                """, (
                    funcionario,
                    COLECTOR_PREDETERMINADO,
                    ahora,
                    ahora,
                    ahora
                ))

                conexion.execute("""
                    INSERT INTO periodos_actividad (actividad_id, inicio)
                    VALUES (?, ?)
                """, (cursor.lastrowid, ahora))

            elif int(registro["activa"]) == 1:
                tiempo_acumulado = calcular_tiempo(registro, ahora)

                cerrar_periodo_activo(conexion, registro["id"], ahora)

                conexion.execute("""
                    UPDATE actividades
                    SET tiempo_acumulado = ?,
                        activa = 0,
                        ultimo_inicio = NULL,
                        ultimo_ping = ?
                    WHERE id = ?
                """, (
                    tiempo_acumulado,
                    ahora,
                    registro["id"]
                ))

            else:
                conexion.execute("""
                    UPDATE actividades
                    SET activa = 1,
                        ultimo_inicio = ?,
                        ultimo_ping = ?
                    WHERE id = ?
                """, (
                    ahora,
                    ahora,
                    registro["id"]
                ))

                conexion.execute("""
                    INSERT INTO periodos_actividad (actividad_id, inicio)
                    VALUES (?, ?)
                """, (registro["id"], ahora))

            registro = obtener_registro_abierto(conexion, funcionario)
            respuesta = serializar_registro(registro, ahora)
            respuesta["tiene_registro"] = True

            return jsonify(respuesta)
    finally:
        conexion.close()


@app.route("/api/handheld/controles")
def api_handheld_controles():
    if "handheld_usuario" not in session:
        return jsonify({"error": "Sesion no valida"}), 401

    conexion = conectar_db()
    try:
        controles = conexion.execute("""
            SELECT controles_embarque.id, embarques.proveedor, embarques.barco,
                   embarques.fecha, colectores.numero AS colector
            FROM controles_embarque
            INNER JOIN embarques ON embarques.id = controles_embarque.embarque_id
            INNER JOIN usuarios ON usuarios.id = controles_embarque.usuario_id
            LEFT JOIN colectores ON colectores.id = controles_embarque.colector_id
            WHERE controles_embarque.estado = 'En control'
              AND usuarios.usuario = ?
            ORDER BY embarques.fecha ASC, controles_embarque.id DESC
        """, (session.get("handheld_usuario"),)).fetchall()
        return jsonify({"controles": [dict(control) for control in controles]})
    finally:
        conexion.close()


@app.route("/api/handheld/controles/<int:control_id>")
def api_handheld_control_detalle(control_id):
    if "handheld_usuario" not in session:
        return jsonify({"error": "Sesion no valida"}), 401

    conexion = conectar_db()
    try:
        control = conexion.execute("""
            SELECT controles_embarque.id, controles_embarque.estado,
                   embarques.id AS embarque_id, embarques.proveedor, embarques.barco,
                   colectores.numero AS colector
            FROM controles_embarque
            INNER JOIN embarques ON embarques.id = controles_embarque.embarque_id
            INNER JOIN usuarios ON usuarios.id = controles_embarque.usuario_id
            LEFT JOIN colectores ON colectores.id = controles_embarque.colector_id
            WHERE controles_embarque.id = ? AND usuarios.usuario = ?
        """, (control_id, session.get("handheld_usuario"))).fetchone()
        if control is None:
            return jsonify({"error": "Control no encontrado"}), 404
        articulos = conexion.execute("""
            SELECT sku, descripcion, cantidad_programada, cantidad_controlada
            FROM articulos_embarque
            WHERE embarque_id = ? ORDER BY sku
        """, (control["embarque_id"],)).fetchall()
        return jsonify({
            "control": dict(control),
            "articulos": [{
                "sku": fila["sku"], "descripcion": fila["descripcion"],
                "programada": fila["cantidad_programada"],
                "controlada": fila["cantidad_controlada"],
                "diferencia": fila["cantidad_controlada"] - fila["cantidad_programada"]
            } for fila in articulos]
        })
    finally:
        conexion.close()


@app.route("/api/handheld/controles/<int:control_id>/registrar", methods=["POST"])
def api_handheld_control_registrar(control_id):
    if "handheld_usuario" not in session:
        return jsonify({"ok": False, "mensaje": "Sesión no válida"}), 401

    datos = request.get_json(silent=True) or request.form
    sku = str(datos.get("sku", "")).strip()
    try:
        cantidad = float(datos.get("cantidad", 0))
    except (TypeError, ValueError):
        cantidad = 0
    if not sku or cantidad <= 0:
        return jsonify({"ok": False, "mensaje": "Ingresa un SKU y una cantidad válida."}), 400

    conexion = conectar_db()
    try:
        control = conexion.execute("""
            SELECT controles_embarque.id, controles_embarque.embarque_id
            FROM controles_embarque
            INNER JOIN usuarios ON usuarios.id = controles_embarque.usuario_id
            WHERE controles_embarque.id = ? AND controles_embarque.estado = 'En control'
              AND usuarios.usuario = ?
        """, (control_id, session.get("handheld_usuario"))).fetchone()
        if control is None:
            return jsonify({"ok": False, "mensaje": "El control no está activo."}), 409
        articulo = conexion.execute("""
            SELECT id FROM articulos_embarque WHERE embarque_id = ? AND sku = ?
        """, (control["embarque_id"], sku)).fetchone()
        if articulo is None:
            return jsonify({"ok": False, "mensaje": "El SKU no pertenece al control seleccionado."}), 404
        ahora = time.time()
        with conexion:
            conexion.execute("""
                INSERT INTO movimientos_control_embarque
                (control_id, articulo_id, cantidad, registrado_en)
                VALUES (?, ?, ?, ?)
            """, (control["id"], articulo["id"], cantidad, ahora))
            conexion.execute("""
                UPDATE articulos_embarque
                SET cantidad_controlada = cantidad_controlada + ? WHERE id = ?
            """, (cantidad, articulo["id"]))
            conexion.execute("""
                UPDATE controles_embarque SET ultimo_ping = ? WHERE id = ?
            """, (ahora, control["id"]))
        return jsonify({"ok": True, "mensaje": "Cantidad registrada."})
    finally:
        conexion.close()


@app.route("/actividad/exportar")
def exportar_actividad():
    if "usuario" not in session:
        return redirect("/")

    fecha = request.args.get("fecha", "").strip()
    limites = limites_de_fecha(fecha) if fecha else None

    if fecha and limites is None:
        return redirect("/actividad")

    ahora = datetime.now().timestamp()
    conexion = conectar_db()

    try:
        if limites is not None:
            inicio_dia, fin_dia = limites
            periodos = conexion.execute("""
                SELECT
                    actividades.funcionario,
                    actividades.colector,
                    periodos_actividad.inicio,
                    periodos_actividad.fin
                FROM periodos_actividad
                INNER JOIN actividades
                    ON actividades.id = periodos_actividad.actividad_id
                WHERE periodos_actividad.inicio < ?
                  AND COALESCE(periodos_actividad.fin, ?) > ?
                ORDER BY periodos_actividad.inicio ASC
            """, (fin_dia, ahora, inicio_dia)).fetchall()
        else:
            periodos = conexion.execute("""
                SELECT
                    actividades.funcionario,
                    actividades.colector,
                    periodos_actividad.inicio,
                    periodos_actividad.fin
                FROM periodos_actividad
                INNER JOIN actividades
                    ON actividades.id = periodos_actividad.actividad_id
                ORDER BY periodos_actividad.inicio ASC
            """).fetchall()

        libro = Workbook()
        hoja = libro.active
        hoja.title = "Actividad"
        encabezados = [
            "Fecha", "Funcionario", "Status", "Período activo",
            "Tiempo activo", "Colector"
        ]
        hoja.append(encabezados)

        relleno = PatternFill("solid", fgColor="315F36")

        for celda in hoja[1]:
            celda.fill = relleno
            celda.font = Font(color="FFFFFF", bold=True)
            celda.alignment = Alignment(horizontal="center")

        for periodo in periodos:
            inicio_periodo = float(periodo["inicio"])
            fin_periodo = float(periodo["fin"] or ahora)

            if limites is not None:
                inicio_periodo = max(inicio_periodo, inicio_dia)
                fin_periodo = min(fin_periodo, fin_dia)

            segundos = max(0, int(fin_periodo - inicio_periodo))
            tramo = (
                f"{datetime.fromtimestamp(inicio_periodo).strftime('%H:%M')} - "
                f"{datetime.fromtimestamp(fin_periodo).strftime('%H:%M')}"
            )

            hoja.append([
                fecha if fecha else fecha_desde_marca(inicio_periodo),
                periodo["funcionario"],
                "Activo",
                tramo,
                formatear_tiempo(segundos),
                periodo["colector"]
            ])

        anchos = [14, 28, 14, 22, 18, 22]

        for columna, ancho in enumerate(anchos, start=1):
            hoja.column_dimensions[chr(64 + columna)].width = ancho

        hoja.freeze_panes = "A2"

        archivo = BytesIO()
        libro.save(archivo)
        archivo.seek(0)

        return send_file(
            archivo,
            as_attachment=True,
            download_name=(
                f"actividad_{fecha}.xlsx"
                if fecha else "actividad_completa.xlsx"
            ),
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )
    finally:
        conexion.close()


@app.route("/actividad/exportar-filtrado")
def exportar_actividad_filtrada():
    if "usuario" not in session:
        return redirect("/")

    filtros = {
        "fecha": request.args.get("fecha", "").strip(),
        "fecha_desde": request.args.get("fecha_desde", "").strip(),
        "fecha_hasta": request.args.get("fecha_hasta", "").strip(),
        "funcionario": request.args.get("funcionario", "").strip(),
        "consulta": request.args.get("consulta", "").strip()
    }
    ahora = time.time()
    conexion = conectar_db()
    try:
        registros = obtener_registros_actividad_filtrados(conexion, ahora=ahora, **filtros)
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Actividad"
        hoja.append(["Fecha", "Funcionario", "Status", "Tiempo activo", "Colector"])
        relleno = PatternFill("solid", fgColor="315F36")
        for celda in hoja[1]:
            celda.fill = relleno
            celda.font = Font(color="FFFFFF", bold=True)
            celda.alignment = Alignment(horizontal="center")
        for registro in registros:
            dato = serializar_registro(registro, ahora)
            hoja.append([
                dato["fecha"], dato["funcionario"], dato["status"],
                dato["tiempo"], dato["colector"]
            ])
        for letra, ancho in zip("ABCDE", [14, 28, 25, 18, 24]):
            hoja.column_dimensions[letra].width = ancho
        hoja.freeze_panes = "A2"
        archivo = BytesIO()
        libro.save(archivo)
        archivo.seek(0)
        return send_file(
            archivo, as_attachment=True, download_name="actividad_filtrada.xlsx",
            mimetype=("application/vnd.openxmlformats-officedocument."
                      "spreadsheetml.sheet")
        )
    finally:
        conexion.close()


def obtener_registro_abierto_por_colector(conexion, funcionario, colector):
    return conexion.execute("""
        SELECT *
        FROM actividades
        WHERE funcionario = ?
          AND colector = ?
          AND estado != 'Actividad finalizada'
        ORDER BY id DESC
        LIMIT 1
    """, (funcionario, colector)).fetchone()


def respuesta_actividad_android(funcionario, colector):
    conexion = conectar_db()

    try:
        ahora = time.time()

        with conexion:
            cerrar_registros_inactivos(conexion, ahora)
            registro = obtener_registro_abierto_por_colector(
                conexion, funcionario, colector
            )

            if registro is None:
                return jsonify({
                    "tiene_registro": False,
                    "activa": False,
                    "tiempo_segundos": 0,
                    "tiempo": "00:00:00",
                    "colector": colector
                })

            conexion.execute("""
                UPDATE actividades
                SET ultimo_ping = ?
                WHERE id = ?
            """, (ahora, registro["id"]))

            registro = obtener_registro_abierto_por_colector(
                conexion, funcionario, colector
            )
            respuesta = serializar_registro(registro, ahora)
            respuesta["tiene_registro"] = True
            return jsonify(respuesta)
    finally:
        conexion.close()


@app.route("/api/android/login", methods=["POST"])
def api_android_login():
    datos = request.get_json(silent=True) or {}
    usuario = str(datos.get("usuario", "")).strip()
    contrasena = str(datos.get("contrasena", ""))
    colector = str(datos.get("colector", "")).strip()
    clave_instalacion = str(datos.get("clave_instalacion", "")).strip()
    dispositivo_id = str(datos.get("dispositivo_id", "")).strip()

    if not usuario or not contrasena or not colector or not clave_instalacion or not dispositivo_id:
        return jsonify({"error": "Completa todos los datos de acceso."}), 400

    if not credenciales_correctas(usuario, contrasena):
        return jsonify({"error": "Credenciales erróneas."}), 401

    registro_usuario = obtener_usuario(usuario)
    if registro_usuario is None or bool(registro_usuario["es_admin"]):
        return jsonify({"error": "Las cuentas Administrador solo acceden al panel web."}), 403

    conexion = conectar_db()

    try:
        equipo = conexion.execute("""
            SELECT numero, clave_instalacion_hash, dispositivo_id
            FROM colectores
            WHERE numero = ? AND activo = 1
        """, (colector,)).fetchone()

        if (
            equipo is None
            or not equipo["clave_instalacion_hash"]
            or not hmac.compare_digest(
                equipo["clave_instalacion_hash"], hash_token(clave_instalacion)
            )
        ):
            return jsonify({"error": "Handheld no autorizado."}), 403
        if equipo["dispositivo_id"] and not hmac.compare_digest(equipo["dispositivo_id"], dispositivo_id):
            return jsonify({"error": "Este colector ya está asociado a otro dispositivo."}), 403

        token = secrets.token_urlsafe(32)
        ahora = time.time()

        with conexion:
            # Un handheld representa un único equipo: una nueva sesión anula
            # el acceso anterior de ese mismo colector.
            conexion.execute("""
                UPDATE sesiones_handheld
                SET revocada = 1
                WHERE colector = ?
            """, (equipo["numero"],))
            conexion.execute("""
                INSERT INTO sesiones_handheld (
                    token_hash, funcionario, colector, creado_en, ultimo_ping
                ) VALUES (?, ?, ?, ?, ?)
            """, (
                hash_token(token), usuario, equipo["numero"], ahora, ahora
            ))
            conexion.execute("""
                UPDATE colectores
                SET dispositivo_id = ?, emparejado_en = COALESCE(emparejado_en, ?),
                    ultimo_ping = ?, estado_dispositivo = 'En línea'
                WHERE numero = ?
            """, (dispositivo_id, ahora, ahora, equipo["numero"]))

        return jsonify({
            "token": token,
            "funcionario": usuario,
            "colector": equipo["numero"]
        })
    finally:
        conexion.close()


@app.route("/api/android/actividad")
def api_android_actividad():
    sesion_android = obtener_sesion_android()

    if sesion_android is None:
        return jsonify({"error": "Sesión no válida."}), 401

    return respuesta_actividad_android(
        sesion_android["funcionario"], sesion_android["colector"]
    )


def control_asignado_a_sesion(conexion, control_id, sesion_android):
    return conexion.execute("""
        SELECT controles_embarque.id, controles_embarque.embarque_id,
               controles_embarque.estado, embarques.proveedor, embarques.barco
        FROM controles_embarque
        INNER JOIN colectores ON colectores.id = controles_embarque.colector_id
        INNER JOIN usuarios ON usuarios.id = controles_embarque.usuario_id
        INNER JOIN embarques ON embarques.id = controles_embarque.embarque_id
        WHERE controles_embarque.id = ?
          AND controles_embarque.estado = 'En control'
          AND colectores.numero = ?
          AND usuarios.usuario = ?
    """, (control_id, sesion_android["colector"], sesion_android["funcionario"])).fetchone()


@app.route("/api/android/controles")
def api_android_controles():
    sesion_android = obtener_sesion_android()
    if sesion_android is None:
        return jsonify({"error": "Sesión no válida."}), 401
    conexion = conectar_db()
    try:
        controles = conexion.execute("""
            SELECT controles_embarque.id, embarques.proveedor, embarques.barco,
                   embarques.fecha
            FROM controles_embarque
            INNER JOIN colectores ON colectores.id = controles_embarque.colector_id
            INNER JOIN usuarios ON usuarios.id = controles_embarque.usuario_id
            INNER JOIN embarques ON embarques.id = controles_embarque.embarque_id
            WHERE controles_embarque.estado = 'En control'
              AND colectores.numero = ? AND usuarios.usuario = ?
            ORDER BY embarques.fecha, controles_embarque.id DESC
        """, (sesion_android["colector"], sesion_android["funcionario"])).fetchall()
        return jsonify({"controles": [dict(fila) for fila in controles]})
    finally:
        conexion.close()


@app.route("/api/android/controles/<int:control_id>")
def api_android_control_detalle(control_id):
    sesion_android = obtener_sesion_android()
    if sesion_android is None:
        return jsonify({"error": "Sesión no válida."}), 401
    conexion = conectar_db()
    try:
        control = control_asignado_a_sesion(conexion, control_id, sesion_android)
        if control is None:
            return jsonify({"error": "Control no asignado a esta cuenta y colector."}), 403
        articulos = conexion.execute("""
            SELECT sku, descripcion, cantidad_programada, cantidad_controlada
            FROM articulos_embarque WHERE embarque_id = ? ORDER BY sku
        """, (control["embarque_id"],)).fetchall()
        return jsonify({"control": dict(control), "articulos": [dict(fila) for fila in articulos]})
    finally:
        conexion.close()


@app.route("/api/android/controles/<int:control_id>/registrar", methods=["POST"])
def api_android_control_registrar(control_id):
    sesion_android = obtener_sesion_android()
    if sesion_android is None:
        return jsonify({"error": "Sesión no válida."}), 401
    datos = request.get_json(silent=True) or {}
    sku = str(datos.get("sku", "")).strip()
    try:
        cantidad = float(datos.get("cantidad", 0))
    except (TypeError, ValueError):
        cantidad = 0
    if not sku or cantidad <= 0:
        return jsonify({"error": "Indica SKU y una cantidad válida."}), 400
    conexion = conectar_db()
    try:
        control = control_asignado_a_sesion(conexion, control_id, sesion_android)
        if control is None:
            return jsonify({"error": "Control no autorizado."}), 403
        articulo = conexion.execute("""
            SELECT id FROM articulos_embarque WHERE embarque_id = ? AND sku = ?
        """, (control["embarque_id"], sku)).fetchone()
        if articulo is None:
            return jsonify({"error": "El SKU no pertenece a este embarque."}), 404
        ahora = time.time()
        with conexion:
            conexion.execute("""
                INSERT INTO movimientos_control_embarque
                (control_id, articulo_id, cantidad, registrado_en)
                VALUES (?, ?, ?, ?)
            """, (control_id, articulo["id"], cantidad, ahora))
            conexion.execute("""
                UPDATE articulos_embarque
                SET cantidad_controlada = cantidad_controlada + ? WHERE id = ?
            """, (cantidad, articulo["id"]))
            conexion.execute("UPDATE controles_embarque SET ultimo_ping = ? WHERE id = ?", (ahora, control_id))
        return jsonify({"ok": True, "mensaje": "Cantidad registrada."})
    finally:
        conexion.close()


@app.route("/api/android/heartbeat", methods=["POST"])
def api_android_heartbeat():
    """Telemetría ligera del equipo; no modifica el tiempo operativo."""
    sesion_android = obtener_sesion_android()
    if sesion_android is None:
        return jsonify({"error": "Sesión no válida."}), 401
    datos = request.get_json(silent=True) or {}
    ahora = time.time()
    try:
        bateria = max(0, min(100, int(datos.get("bateria", -1))))
    except (TypeError, ValueError):
        bateria = -1
    try:
        movimiento = float(datos.get("ultimo_movimiento", ahora))
        if movimiento > 100000000000:
            movimiento /= 1000
    except (TypeError, ValueError):
        movimiento = ahora
    # Evita que una marca errónea de reloj altere el registro de actividad.
    movimiento = min(ahora, max(ahora - 7200, movimiento))
    cargando = 1 if bool(datos.get("cargando", False)) else 0
    conexion = conectar_db()
    try:
        with conexion:
            conexion.execute("""
                UPDATE colectores
                SET ultimo_ping = ?, ultimo_movimiento = ?,
                    bateria_porcentaje = CASE WHEN ? >= 0 THEN ? ELSE bateria_porcentaje END,
                    cargando = ?, estado_dispositivo = 'En línea'
                WHERE numero = ?
            """, (ahora, movimiento, bateria, bateria, cargando, sesion_android["colector"]))
        return jsonify({"ok": True, "servidor_en": ahora})
    finally:
        conexion.close()


@app.route("/api/android/inactividad", methods=["POST"])
def api_android_inactividad():
    """Cierra desde el último movimiento válido, sin sumar el lapso inmóvil."""
    sesion_android = obtener_sesion_android()
    if sesion_android is None:
        return jsonify({"error": "Sesión no válida."}), 401
    datos = request.get_json(silent=True) or {}
    ahora = time.time()
    try:
        ultimo_movimiento = float(datos.get("ultimo_movimiento", ahora))
        if ultimo_movimiento > 100000000000:
            ultimo_movimiento /= 1000
    except (TypeError, ValueError):
        ultimo_movimiento = ahora
    ultimo_movimiento = min(ahora, max(ahora - 7200, ultimo_movimiento))
    conexion = conectar_db()
    try:
        with conexion:
            registro = obtener_registro_abierto_por_colector(
                conexion, sesion_android["funcionario"], sesion_android["colector"]
            )
            if registro is not None:
                cierre = max(float(registro["ultimo_inicio"] or registro["inicio"]), ultimo_movimiento)
                finalizar_registro(conexion, registro, cierre)
            conexion.execute("""
                UPDATE sesiones_handheld SET revocada = 1 WHERE id = ?
            """, (sesion_android["id"],))
            conexion.execute("""
                UPDATE colectores
                SET ultimo_movimiento = ?, ultimo_ping = ?,
                    estado_dispositivo = 'Sesión finalizada por inactividad'
                WHERE numero = ?
            """, (ultimo_movimiento, ahora, sesion_android["colector"]))
        return jsonify({"ok": True, "motivo": "Inactividad prolongada"})
    finally:
        conexion.close()


def funcionario_en_horario(horario, hora_actual):
    coincidencia = re.fullmatch(r"\s*(\d{1,2}:\d{2})\s*(?:-|a)\s*(\d{1,2}:\d{2})\s*", horario or "", re.IGNORECASE)
    if coincidencia is None:
        return False
    try:
        inicio = datetime.strptime(coincidencia.group(1), "%H:%M").time()
        fin = datetime.strptime(coincidencia.group(2), "%H:%M").time()
    except ValueError:
        return False
    return inicio <= hora_actual < fin if inicio <= fin else (hora_actual >= inicio or hora_actual < fin)


@app.route("/api/panel/disponibilidad")
@admin_requerido
def disponibilidad_personal():
    ahora = datetime.now()
    fecha = ahora.date().isoformat()
    codigo_empresa = empresa_actual()
    conexion = conectar_db()
    try:
        condicion_empresa = ""
        parametros = []
        if codigo_empresa:
            condicion_empresa = """
                INNER JOIN usuario_empresas ON usuario_empresas.usuario_id = usuarios.id
                    AND usuario_empresas.empresa_codigo = ?
            """
            parametros.append(codigo_empresa)
        funcionarios = conexion.execute(f"""
            SELECT DISTINCT usuarios.id,
                   COALESCE(NULLIF(usuarios.nombre_funcionario, ''), usuarios.usuario) AS funcionario
            FROM usuarios
            {condicion_empresa}
            WHERE usuarios.es_admin = 0 AND usuarios.activo = 1
            ORDER BY funcionario
        """, parametros).fetchall()
        disponibles = []
        no_disponibles = []
        for funcionario in funcionarios:
            ausencia = conexion.execute("""
                SELECT tipo FROM ausencias_funcionarios
                WHERE usuario_id = ? AND fecha_desde <= ? AND fecha_hasta >= ?
                LIMIT 1
            """, (funcionario["id"], fecha, fecha)).fetchone()
            if ausencia:
                no_disponibles.append({"funcionario": funcionario["funcionario"], "motivo": ausencia["tipo"]})
                continue
            condiciones_plan = ""
            parametros_turno = [funcionario["id"], fecha, session.get("usuario")]
            if codigo_empresa:
                condiciones_plan = " AND planificaciones_horarios.empresa_codigo = ?"
                parametros_turno.append(codigo_empresa)
            turno = conexion.execute(f"""
                SELECT turnos_personal.horario
                FROM turnos_personal
                INNER JOIN planificaciones_horarios
                    ON planificaciones_horarios.id = turnos_personal.planificacion_id
                WHERE turnos_personal.usuario_id = ? AND turnos_personal.fecha = ?
                  AND planificaciones_horarios.supervisor_usuario = ?
                  {condiciones_plan}
                ORDER BY planificaciones_horarios.actualizado_en DESC
                LIMIT 1
            """, parametros_turno).fetchone()
            horario = turno["horario"] if turno else ""
            if funcionario_en_horario(horario, ahora.time()):
                disponibles.append(funcionario["funcionario"])
            else:
                no_disponibles.append({"funcionario": funcionario["funcionario"], "motivo": horario or "Sin turno"})
        return jsonify({
            "disponibles": len(disponibles), "total": len(funcionarios),
            "nombres": disponibles, "no_disponibles": no_disponibles
        })
    finally:
        conexion.close()


@app.route("/api/panel/tareas-hoy")
@admin_requerido
def tareas_hoy_panel():
    fecha = datetime.now().date().isoformat()
    conexion = conectar_db()
    try:
        tareas = conexion.execute("""
            SELECT id, titulo, hora, prioridad, estado
            FROM tareas_supervision
            WHERE supervisor_usuario = ? AND fecha = ? AND estado != 'Completada'
            ORDER BY CASE prioridad WHEN 'Alta' THEN 1 WHEN 'Media' THEN 2 ELSE 3 END,
                     COALESCE(hora, '') ASC, id ASC
            LIMIT 6
        """, (session.get("usuario"), fecha)).fetchall()
        return jsonify({"fecha": fecha, "tareas": [dict(tarea) for tarea in tareas]})
    finally:
        conexion.close()


@app.route("/api/android/actividad/toggle", methods=["POST"])
def api_android_toggle_actividad():
    sesion_android = obtener_sesion_android()

    if sesion_android is None:
        return jsonify({"error": "Sesión no válida."}), 401

    funcionario = sesion_android["funcionario"]
    colector = sesion_android["colector"]
    conexion = conectar_db()

    try:
        ahora = time.time()

        with conexion:
            cerrar_registros_inactivos(conexion, ahora)
            registro = obtener_registro_abierto_por_colector(
                conexion, funcionario, colector
            )

            if registro is None:
                cursor = conexion.execute("""
                    INSERT INTO actividades (
                        funcionario, colector, estado, inicio, ultimo_inicio,
                        tiempo_acumulado, activa, ultimo_ping
                    ) VALUES (?, ?, 'En operacion', ?, ?, 0, 1, ?)
                """, (funcionario, colector, ahora, ahora, ahora))
                conexion.execute("""
                    INSERT INTO periodos_actividad (actividad_id, inicio)
                    VALUES (?, ?)
                """, (cursor.lastrowid, ahora))

            elif int(registro["activa"]) == 1:
                tiempo_acumulado = calcular_tiempo(registro, ahora)
                cerrar_periodo_activo(conexion, registro["id"], ahora)
                conexion.execute("""
                    UPDATE actividades
                    SET tiempo_acumulado = ?, activa = 0,
                        ultimo_inicio = NULL, ultimo_ping = ?
                    WHERE id = ?
                """, (tiempo_acumulado, ahora, registro["id"]))

            else:
                conexion.execute("""
                    UPDATE actividades
                    SET activa = 1, ultimo_inicio = ?, ultimo_ping = ?
                    WHERE id = ?
                """, (ahora, ahora, registro["id"]))
                conexion.execute("""
                    INSERT INTO periodos_actividad (actividad_id, inicio)
                    VALUES (?, ?)
                """, (registro["id"], ahora))

        return respuesta_actividad_android(funcionario, colector)
    finally:
        conexion.close()


@app.route("/api/android/logout", methods=["POST"])
def api_android_logout():
    sesion_android = obtener_sesion_android()

    if sesion_android is None:
        return jsonify({"error": "Sesión no válida."}), 401

    conexion = conectar_db()

    try:
        ahora = time.time()

        with conexion:
            cerrar_registros_inactivos(conexion, ahora)
            registro = obtener_registro_abierto_por_colector(
                conexion,
                sesion_android["funcionario"],
                sesion_android["colector"]
            )

            if registro is not None:
                finalizar_registro(conexion, registro, ahora)

            conexion.execute("""
                UPDATE sesiones_handheld
                SET revocada = 1
                WHERE id = ?
            """, (sesion_android["id"],))

        return jsonify({"ok": True})
    finally:
        conexion.close()


@app.route("/colectores/regenerar-clave", methods=["POST"])
@admin_requerido
def regenerar_clave_colector():
    identificador = request.form.get("id", "")
    conexion = conectar_db()

    try:
        with conexion:
            resultado = conexion.execute("""
                UPDATE colectores
                SET clave_instalacion = NULL, clave_instalacion_hash = NULL,
                    dispositivo_id = NULL, emparejado_en = NULL,
                    estado_dispositivo = 'Pendiente de enlace'
                WHERE id = ?
            """, (identificador,))

            conexion.execute("""
                UPDATE sesiones_handheld
                SET revocada = 1
                WHERE colector = (
                    SELECT numero FROM colectores WHERE id = ?
                )
            """, (identificador,))

        if resultado.rowcount == 0:
            return redirigir_usuarios("No se encontró el handheld seleccionado.")

        return redirigir_usuarios(
            "Clave de instalación renovada. Configura nuevamente la app."
        )
    finally:
        conexion.close()


@app.route("/api/actividad/registros")
def api_registros_actividad():
    if "usuario" not in session:
        return jsonify({"error": "Sesion no valida"}), 401

    conexion = conectar_db()

    try:
        ahora = time.time()

        with conexion:
            cerrar_registros_inactivos(conexion, ahora)

            registros = obtener_registros_actividad_filtrados(
                conexion,
                request.args.get("fecha", "").strip(),
                request.args.get("fecha_desde", "").strip(),
                request.args.get("fecha_hasta", "").strip(),
                request.args.get("funcionario", "").strip(),
                request.args.get("consulta", "").strip(),
                ahora
            )

            datos = [
                serializar_registro(registro, ahora)
                for registro in registros
            ]

            hay_operacion = any(
                registro["status"].startswith("En operacion")
                for registro in datos
            )

            return jsonify({
                "resumen": "En operacion" if hay_operacion else "Sin actividad",
                "registros": datos
            })
    finally:
        conexion.close()


if __name__ == "__main__":
    iniciar_monitor_de_inactividad()

    # La app queda disponible únicamente en la red local de la PC. No se usa
    # el modo depuración para no exponer el depurador de Flask al Wi-Fi.
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5000")), debug=False)
