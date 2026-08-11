from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    session,
    jsonify,
    send_file
)
from openpyxl import load_workbook, Workbook
from urllib.parse import quote
from io import BytesIO
from datetime import datetime
from database import conectar
import os
import time
import unicodedata

embarques_bp = Blueprint("embarques", __name__)
RUTA_BD = os.environ.get(
    "POLO_OESTE_DB",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "actividad.db")
)


@embarques_bp.app_template_filter("fecha_hora")
def fecha_hora(marca_tiempo):
    if not marca_tiempo:
        return "Sin registro"

    return datetime.fromtimestamp(float(marca_tiempo)).strftime("%d/%m/%Y %H:%M")


def db():
    return conectar(RUTA_BD)


def requerido():
    return "usuario" in session


def empresa_actual():
    """Operativa elegida en la barra global; vacÃ­a significa ver todas."""
    return str(session.get("empresa_codigo", "")).strip()


def mensaje(texto):
    return redirect("/embarques?mensaje=" + quote(texto))


def normalizar(texto):
    texto = str(texto or "").strip().lower()

    return "".join(
        letra for letra in unicodedata.normalize("NFD", texto)
        if unicodedata.category(letra) != "Mn"
    )


def crear_tablas():
    conexion = db()

    try:
        conexion.execute("""
            CREATE TABLE IF NOT EXISTS embarques (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT NOT NULL,
                proveedor TEXT NOT NULL,
                barco TEXT NOT NULL,
                bultos INTEGER NOT NULL,
                descripcion TEXT NOT NULL,
                estado TEXT NOT NULL DEFAULT 'Proximo',
                confirmado_en REAL,
                finalizado_en REAL,
                empresa_codigo TEXT
            )
        """)

        columnas_embarques = {
            columna["name"]
            for columna in conexion.execute("PRAGMA table_info(embarques)")
        }

        if "confirmado_en" not in columnas_embarques:
            conexion.execute("""
                ALTER TABLE embarques
                ADD COLUMN confirmado_en REAL
            """)

        if "finalizado_en" not in columnas_embarques:
            conexion.execute("""
                ALTER TABLE embarques
                ADD COLUMN finalizado_en REAL
            """)

        if "empresa_codigo" not in columnas_embarques:
            conexion.execute("""
                ALTER TABLE embarques
                ADD COLUMN empresa_codigo TEXT
            """)

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS articulos_embarque (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                embarque_id INTEGER NOT NULL,
                sku TEXT NOT NULL,
                descripcion TEXT NOT NULL,
                cantidad_programada REAL NOT NULL DEFAULT 1,
                cantidad_controlada REAL NOT NULL DEFAULT 0,
                FOREIGN KEY (embarque_id) REFERENCES embarques(id)
            )
        """)

        columnas_articulos = {
            columna["name"]
            for columna in conexion.execute("PRAGMA table_info(articulos_embarque)")
        }
        if "cantidad_programada" not in columnas_articulos:
            conexion.execute("""
                ALTER TABLE articulos_embarque
                ADD COLUMN cantidad_programada REAL NOT NULL DEFAULT 1
            """)
        if "cantidad_controlada" not in columnas_articulos:
            conexion.execute("""
                ALTER TABLE articulos_embarque
                ADD COLUMN cantidad_controlada REAL NOT NULL DEFAULT 0
            """)

        conexion.execute("""
            CREATE TABLE IF NOT EXISTS controles_embarque (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                embarque_id INTEGER NOT NULL UNIQUE,
                colector_id INTEGER,
                usuario_id INTEGER,
                estado TEXT NOT NULL DEFAULT 'Pendiente',
                iniciado_en REAL,
                ultimo_ping REAL,
                finalizado_en REAL,
                FOREIGN KEY (embarque_id) REFERENCES embarques(id)
            )
        """)
        columnas_controles = {
            columna["name"] for columna in conexion.execute("PRAGMA table_info(controles_embarque)")
        }
        if "usuario_id" not in columnas_controles:
            conexion.execute("ALTER TABLE controles_embarque ADD COLUMN usuario_id INTEGER")
        conexion.execute("""
            CREATE TABLE IF NOT EXISTS movimientos_control_embarque (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                control_id INTEGER NOT NULL,
                articulo_id INTEGER NOT NULL,
                cantidad REAL NOT NULL,
                registrado_en REAL NOT NULL,
                FOREIGN KEY (control_id) REFERENCES controles_embarque(id),
                FOREIGN KEY (articulo_id) REFERENCES articulos_embarque(id)
            )
        """)
        conexion.execute("""
            CREATE INDEX IF NOT EXISTS indice_control_embarque_articulo
            ON movimientos_control_embarque(control_id, articulo_id)
        """)

        conexion.commit()
    finally:
        conexion.close()


@embarques_bp.route("/embarques")
def embarques():
    if not requerido():
        return redirect("/")

    conexion = db()

    try:
        codigo_empresa = empresa_actual()
        registros = conexion.execute("""
            SELECT *
            FROM embarques
            WHERE (? = '' OR empresa_codigo = ?)
            ORDER BY fecha DESC, id DESC
        """, (codigo_empresa, codigo_empresa)).fetchall()

        return render_template("embarques.html", embarques=registros)
    finally:
        conexion.close()


@embarques_bp.route("/embarques/exportar")
def exportar_embarques():
    if not requerido():
        return redirect("/")

    fecha = request.args.get("fecha", "").strip()
    fecha_desde = request.args.get("fecha_desde", "").strip()
    fecha_hasta = request.args.get("fecha_hasta", "").strip()
    consulta = normalizar(request.args.get("consulta", ""))
    conexion = db()
    try:
        condiciones = []
        parametros = []
        if fecha:
            condiciones.append("fecha = ?")
            parametros.append(fecha)
        if fecha_desde:
            condiciones.append("fecha >= ?")
            parametros.append(fecha_desde)
        if fecha_hasta:
            condiciones.append("fecha <= ?")
            parametros.append(fecha_hasta)
        codigo_empresa = empresa_actual()
        if codigo_empresa:
            condiciones.append("empresa_codigo = ?")
            parametros.append(codigo_empresa)
        where = " WHERE " + " AND ".join(condiciones) if condiciones else ""
        registros = conexion.execute(f"""
            SELECT * FROM embarques {where} ORDER BY fecha DESC, id DESC
        """, parametros).fetchall()
        if consulta:
            registros = [
                registro for registro in registros
                if consulta in normalizar(" ".join([
                    registro["fecha"], registro["proveedor"], registro["barco"],
                    str(registro["bultos"]), registro["descripcion"], registro["estado"],
                    fecha_hora(registro["confirmado_en"]), fecha_hora(registro["finalizado_en"])
                ]))
            ]

        libro = Workbook()
        hoja = libro.active
        hoja.title = "Embarques"
        hoja.append([
            "Fecha tentativa", "Proveedor", "Barco", "Bultos", "Descripción",
            "Estado", "Confirmado", "Finalizado", "Empresa"
        ])
        for fila in registros:
            hoja.append([
                fila["fecha"], fila["proveedor"], fila["barco"], fila["bultos"],
                fila["descripcion"], fila["estado"], fecha_hora(fila["confirmado_en"]),
                fecha_hora(fila["finalizado_en"]), fila["empresa_codigo"] or "Sin asignar"
            ])
        for letra, ancho in zip("ABCDEFGHI", [17, 28, 26, 12, 46, 16, 22, 22, 14]):
            hoja.column_dimensions[letra].width = ancho
        hoja.freeze_panes = "A2"
        archivo = BytesIO()
        libro.save(archivo)
        archivo.seek(0)
        return send_file(
            archivo, as_attachment=True, download_name="embarques_filtrados.xlsx",
            mimetype=("application/vnd.openxmlformats-officedocument."
                      "spreadsheetml.sheet")
        )
    finally:
        conexion.close()


@embarques_bp.route("/embarques/crear", methods=["POST"])
def crear_embarque():
    if not requerido():
        return redirect("/")

    fecha = request.form.get("fecha", "")
    proveedor = request.form.get("proveedor", "").strip()
    barco = request.form.get("barco", "").strip()
    descripcion = request.form.get("descripcion", "").strip()

    try:
        bultos = int(request.form.get("bultos", "0"))
    except ValueError:
        bultos = 0

    if not fecha or not proveedor or not barco or not descripcion or bultos <= 0:
        return mensaje("Completa todos los campos del embarque.")

    conexion = db()

    try:
        with conexion:
            conexion.execute("""
                INSERT INTO embarques (
                fecha, proveedor, barco, bultos, descripcion, estado, empresa_codigo
            )
                VALUES (?, ?, ?, ?, ?, 'Proximo', ?)
            """, (fecha, proveedor, barco, bultos, descripcion, empresa_actual()))

        return mensaje("Embarque registrado como próximo.")
    finally:
        conexion.close()


@embarques_bp.route("/embarques/<int:embarque_id>/actualizar", methods=["POST"])
def actualizar_embarque(embarque_id):
    if not requerido():
        return redirect("/")

    fecha = request.form.get("fecha", "").strip()
    proveedor = request.form.get("proveedor", "").strip()
    barco = request.form.get("barco", "").strip()
    descripcion = request.form.get("descripcion", "").strip()
    try:
        bultos = int(request.form.get("bultos", "0"))
    except ValueError:
        bultos = 0
    if not fecha or not proveedor or not barco or not descripcion or bultos <= 0:
        return mensaje("Completa todos los campos del embarque.")

    conexion = db()
    try:
        with conexion:
            resultado = conexion.execute("""
                UPDATE embarques
                SET fecha = ?, proveedor = ?, barco = ?, bultos = ?, descripcion = ?
                WHERE id = ?
            """, (fecha, proveedor, barco, bultos, descripcion, embarque_id))
        if not resultado.rowcount:
            return mensaje("No se encontró el embarque seleccionado.")
        return mensaje("Embarque actualizado correctamente.")
    finally:
        conexion.close()


@embarques_bp.route("/embarques/<int:embarque_id>/confirmar", methods=["POST"])
def confirmar_embarque(embarque_id):
    if not requerido():
        return redirect("/")

    archivo = request.files.get("archivo")

    if archivo is None or not archivo.filename.lower().endswith(".xlsx"):
        return mensaje("Selecciona un archivo Excel .xlsx.")

    try:
        libro = load_workbook(archivo, read_only=True, data_only=True)
        hoja = libro.active

        encabezados = [
            normalizar(celda.value)
            for celda in next(hoja.iter_rows(min_row=1, max_row=1))
        ]

        columnas_validas = {"sku", "descripcion", "cantidad"}
        if set(encabezados) == {"sku", "descripcion", "unidades"}:
            encabezados[encabezados.index("unidades")] = "cantidad"
        if set(encabezados) != columnas_validas:
            return mensaje(
                "El Excel debe contener SKU, Descripcion y Cantidad."
            )

        columna_sku = encabezados.index("sku")
        columna_descripcion = encabezados.index("descripcion")
        columna_cantidad = encabezados.index("cantidad")
        articulos = []

        for numero_fila, fila in enumerate(hoja.iter_rows(min_row=2), start=2):
            sku = str(fila[columna_sku].value or "").strip()
            descripcion = str(fila[columna_descripcion].value or "").strip()
            try:
                cantidad = float(fila[columna_cantidad].value or 0)
            except (TypeError, ValueError):
                cantidad = 0

            if not sku and not descripcion and not cantidad:
                continue

            if not sku or not descripcion or cantidad <= 0:
                return mensaje(
                    f"La fila {numero_fila} debe tener SKU, Descripcion y Cantidad válida."
                )

            articulos.append((sku, descripcion, cantidad))

        if not articulos:
            return mensaje("El Excel no contiene artículos válidos.")

    except Exception:
        return mensaje("No fue posible leer el archivo Excel.")

    conexion = db()

    try:
        embarque = conexion.execute("""
            SELECT * FROM embarques WHERE id = ?
        """, (embarque_id,)).fetchone()

        if embarque is None:
            return mensaje("No se encontró el embarque.")

        if embarque["estado"] != "Proximo":
            return mensaje("Este embarque ya fue confirmado.")

        with conexion:
            for sku, descripcion, cantidad in articulos:
                conexion.execute("""
                    INSERT INTO articulos_embarque
                    (embarque_id, sku, descripcion, cantidad_programada)
                    VALUES (?, ?, ?, ?)
                """, (embarque_id, sku, descripcion, cantidad))

            conexion.execute("""
                UPDATE embarques
                SET estado = 'En proceso',
                    confirmado_en = ?
                WHERE id = ?
            """, (time.time(), embarque_id))

        return redirect(
            f"/embarques/{embarque_id}/control?mensaje=" + quote(
                f"Embarque confirmado con {len(articulos)} artículos."
            )
        )
    finally:
        conexion.close()


@embarques_bp.route("/embarques/<int:embarque_id>/finalizar", methods=["POST"])
def finalizar_embarque(embarque_id):
    if not requerido():
        return redirect("/")

    conexion = db()

    try:
        with conexion:
            conexion.execute("""
                UPDATE embarques
                SET estado = 'Finalizado',
                    finalizado_en = ?
                WHERE id = ? AND estado = 'En proceso'
            """, (time.time(), embarque_id))

        return mensaje("Embarque finalizado correctamente.")
    finally:
        conexion.close()


@embarques_bp.route("/embarques/<int:embarque_id>/eliminar", methods=["POST"])
def eliminar_embarque(embarque_id):
    if not requerido():
        return redirect("/")

    conexion = db()

    try:
        with conexion:
            conexion.execute("""
                DELETE FROM articulos_embarque
                WHERE embarque_id = ?
            """, (embarque_id,))

            resultado = conexion.execute("""
                DELETE FROM embarques
                WHERE id = ?
            """, (embarque_id,))

        if resultado.rowcount == 0:
            return mensaje("No se encontró el embarque seleccionado.")

        return mensaje("Embarque eliminado correctamente.")
    finally:
        conexion.close()


@embarques_bp.route("/embarques/<int:embarque_id>/detalle")
def detalle_embarque(embarque_id):
    if not requerido():
        return redirect("/")

    conexion = db()

    try:
        embarque = conexion.execute("""
            SELECT * FROM embarques WHERE id = ?
        """, (embarque_id,)).fetchone()

        cantidad_articulos = conexion.execute("""
            SELECT COUNT(*) AS total
            FROM articulos_embarque
            WHERE embarque_id = ?
        """, (embarque_id,)).fetchone()["total"]

        if embarque is None:
            return redirect("/embarques")

        return render_template(
            "detalle_embarque.html",
            embarque=embarque,
            cantidad_articulos=cantidad_articulos
        )
    finally:
        conexion.close()


@embarques_bp.route("/embarques/<int:embarque_id>/articulos/excel")
def descargar_articulos(embarque_id):
    if not requerido():
        return redirect("/")

    conexion = db()

    try:
        embarque = conexion.execute("""
            SELECT * FROM embarques WHERE id = ?
        """, (embarque_id,)).fetchone()

        articulos = conexion.execute("""
            SELECT sku, descripcion, cantidad_programada
            FROM articulos_embarque
            WHERE embarque_id = ?
            ORDER BY sku
        """, (embarque_id,)).fetchall()

        if embarque is None or embarque["estado"] == "Proximo":
            return redirect("/embarques")

        libro = Workbook()
        hoja = libro.active
        hoja.title = "Articulos"

        hoja.append(["SKU", "Descripcion", "Cantidad"])

        for articulo in articulos:
            hoja.append([
                articulo["sku"], articulo["descripcion"],
                articulo["cantidad_programada"]
            ])

        hoja.column_dimensions["A"].width = 25
        hoja.column_dimensions["B"].width = 55
        hoja.column_dimensions["C"].width = 14

        archivo = BytesIO()
        libro.save(archivo)
        archivo.seek(0)

        return send_file(
            archivo,
            as_attachment=True,
            download_name=f"embarque_{embarque_id}_articulos.xlsx",
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )
    finally:
        conexion.close()


def detalle_control(conexion, embarque_id):
    control = conexion.execute("""
        SELECT controles_embarque.*, colectores.numero AS colector_numero,
               colectores.descripcion AS colector_descripcion
        FROM controles_embarque
        LEFT JOIN colectores ON colectores.id = controles_embarque.colector_id
        WHERE embarque_id = ?
    """, (embarque_id,)).fetchone()
    if control is None:
        return None, []

    articulos = conexion.execute("""
        SELECT articulos_embarque.id, articulos_embarque.sku,
               articulos_embarque.descripcion,
               articulos_embarque.cantidad_programada,
               articulos_embarque.cantidad_controlada
        FROM articulos_embarque
        WHERE embarque_id = ?
        ORDER BY articulos_embarque.sku
    """, (embarque_id,)).fetchall()
    return control, articulos


@embarques_bp.route("/embarques/<int:embarque_id>/control")
def control_embarque(embarque_id):
    if not requerido():
        return redirect("/")

    conexion = db()
    try:
        embarque = conexion.execute("""
            SELECT * FROM embarques WHERE id = ?
        """, (embarque_id,)).fetchone()
        if embarque is None or embarque["estado"] == "Proximo":
            return redirect("/embarques")

        control, articulos = detalle_control(conexion, embarque_id)
        colectores = conexion.execute("""
            SELECT id, numero, descripcion
            FROM colectores WHERE activo = 1
            ORDER BY numero
        """).fetchall()
        funcionarios = conexion.execute("""
            SELECT id, usuario,
                   COALESCE(NULLIF(nombre_funcionario, ''), usuario) AS nombre_funcionario
            FROM usuarios
            WHERE activo = 1 AND es_admin = 0
            ORDER BY nombre_funcionario COLLATE NOCASE
        """).fetchall()
        return render_template(
            "control_embarque.html", embarque=embarque, control=control,
            articulos=articulos, colectores=colectores, funcionarios=funcionarios,
            mensaje=request.args.get("mensaje", "")
        )
    finally:
        conexion.close()


@embarques_bp.route("/embarques/<int:embarque_id>/control/iniciar", methods=["POST"])
def iniciar_control_embarque(embarque_id):
    if not requerido():
        return redirect("/")

    try:
        colector_id = int(request.form.get("colector_id", ""))
    except ValueError:
        colector_id = 0
    try:
        usuario_id = int(request.form.get("usuario_id", ""))
    except ValueError:
        usuario_id = 0

    conexion = db()
    try:
        colector = conexion.execute("""
            SELECT id FROM colectores WHERE id = ? AND activo = 1
        """, (colector_id,)).fetchone()
        funcionario = conexion.execute("""
            SELECT id FROM usuarios WHERE id = ? AND activo = 1 AND es_admin = 0
        """, (usuario_id,)).fetchone()
        embarque = conexion.execute("""
            SELECT id FROM embarques WHERE id = ? AND estado != 'Proximo'
        """, (embarque_id,)).fetchone()
        if colector is None or funcionario is None or embarque is None:
            return redirect(f"/embarques/{embarque_id}/control?mensaje=" + quote(
                "Selecciona un colector activo para iniciar el control."
            ))

        ahora = time.time()
        with conexion:
            existente = conexion.execute("""
                SELECT id, estado FROM controles_embarque WHERE embarque_id = ?
            """, (embarque_id,)).fetchone()
            if existente is None:
                conexion.execute("""
                    INSERT INTO controles_embarque
                    (embarque_id, colector_id, usuario_id, estado, iniciado_en, ultimo_ping)
                    VALUES (?, ?, ?, 'En control', ?, ?)
                """, (embarque_id, colector_id, usuario_id, ahora, ahora))
            elif existente["estado"] == "Pendiente":
                conexion.execute("""
                    UPDATE controles_embarque
                    SET colector_id = ?, usuario_id = ?, estado = 'En control', iniciado_en = ?,
                        ultimo_ping = ?, finalizado_en = NULL
                    WHERE id = ?
                """, (colector_id, usuario_id, ahora, ahora, existente["id"]))
            else:
                return redirect(f"/embarques/{embarque_id}/control")

        return redirect(f"/embarques/{embarque_id}/control?mensaje=" + quote(
            "Control iniciado y colector asignado."
        ))
    finally:
        conexion.close()


@embarques_bp.route("/api/embarques/<int:embarque_id>/control")
def api_control_embarque(embarque_id):
    if not requerido():
        return jsonify({"error": "Sesion no valida"}), 401

    conexion = db()
    try:
        control, articulos = detalle_control(conexion, embarque_id)
        if control is None:
            return jsonify({"control": None, "articulos": []})

        programadas = sum(float(articulo["cantidad_programada"] or 0) for articulo in articulos)
        controladas = sum(float(articulo["cantidad_controlada"] or 0) for articulo in articulos)
        movimientos = conexion.execute("""
            SELECT movimientos_control_embarque.id, articulos_embarque.sku,
                   articulos_embarque.descripcion, movimientos_control_embarque.cantidad,
                   movimientos_control_embarque.registrado_en
            FROM movimientos_control_embarque
            INNER JOIN articulos_embarque
                ON articulos_embarque.id = movimientos_control_embarque.articulo_id
            WHERE movimientos_control_embarque.control_id = ?
            ORDER BY movimientos_control_embarque.registrado_en DESC,
                     movimientos_control_embarque.id DESC
            LIMIT 250
        """, (control["id"],)).fetchall()
        return jsonify({
            "control": {
                "estado": control["estado"],
                "colector": (
                    f"{control['colector_numero']} · {control['colector_descripcion']}"
                    if control["colector_numero"] else "Sin asignar"
                ),
                "programadas": programadas,
                "controladas": controladas,
                "porcentaje": round((controladas / programadas) * 100, 1) if programadas else 0
            },
            "articulos": [
                {
                    "sku": articulo["sku"],
                    "descripcion": articulo["descripcion"],
                    "programada": articulo["cantidad_programada"],
                    "controlada": articulo["cantidad_controlada"],
                    "diferencia": articulo["cantidad_controlada"] - articulo["cantidad_programada"]
                }
                for articulo in articulos
            ],
            "movimientos": [
                {
                    "id": movimiento["id"], "sku": movimiento["sku"],
                    "descripcion": movimiento["descripcion"],
                    "cantidad": movimiento["cantidad"],
                    "registrado_en": fecha_hora(movimiento["registrado_en"])
                }
                for movimiento in movimientos
            ]
        })
    finally:
        conexion.close()


@embarques_bp.route("/api/embarques/<int:embarque_id>/control/movimientos/<int:movimiento_id>", methods=["PUT", "DELETE"])
def modificar_movimiento_control(embarque_id, movimiento_id):
    if not requerido():
        return jsonify({"ok": False, "mensaje": "Sesión no válida"}), 401

    conexion = db()
    try:
        control = conexion.execute("""
            SELECT id FROM controles_embarque
            WHERE embarque_id = ? AND estado = 'En control'
        """, (embarque_id,)).fetchone()
        if control is None:
            return jsonify({"ok": False, "mensaje": "El control no está activo."}), 409
        movimiento = conexion.execute("""
            SELECT id, articulo_id, cantidad
            FROM movimientos_control_embarque
            WHERE id = ? AND control_id = ?
        """, (movimiento_id, control["id"])).fetchone()
        if movimiento is None:
            return jsonify({"ok": False, "mensaje": "Registro no encontrado."}), 404

        if request.method == "DELETE":
            with conexion:
                conexion.execute("DELETE FROM movimientos_control_embarque WHERE id = ?", (movimiento_id,))
                conexion.execute("""
                    UPDATE articulos_embarque
                    SET cantidad_controlada = MAX(0, cantidad_controlada - ?)
                    WHERE id = ?
                """, (movimiento["cantidad"], movimiento["articulo_id"]))
            return jsonify({"ok": True, "mensaje": "Registro eliminado."})

        datos = request.get_json(silent=True) or {}
        try:
            cantidad = float(datos.get("cantidad", 0))
        except (TypeError, ValueError):
            cantidad = 0
        if cantidad <= 0:
            return jsonify({"ok": False, "mensaje": "La cantidad debe ser mayor a cero."}), 400
        diferencia = cantidad - float(movimiento["cantidad"])
        with conexion:
            conexion.execute("""
                UPDATE movimientos_control_embarque SET cantidad = ? WHERE id = ?
            """, (cantidad, movimiento_id))
            conexion.execute("""
                UPDATE articulos_embarque
                SET cantidad_controlada = MAX(0, cantidad_controlada + ?)
                WHERE id = ?
            """, (diferencia, movimiento["articulo_id"]))
        return jsonify({"ok": True, "mensaje": "Registro actualizado."})
    finally:
        conexion.close()


@embarques_bp.route("/embarques/<int:embarque_id>/control/exportar")
def exportar_control_embarque(embarque_id):
    if not requerido():
        return redirect("/")

    conexion = db()
    try:
        embarque = conexion.execute("""
            SELECT proveedor, barco FROM embarques WHERE id = ?
        """, (embarque_id,)).fetchone()
        control, articulos = detalle_control(conexion, embarque_id)
        if embarque is None or control is None:
            return redirect("/embarques")
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Control de ingreso"
        hoja.append([
            "SKU", "Descripción", "Cantidad programada", "Cantidad controlada",
            "Diferencia", "Estado"
        ])
        for articulo in articulos:
            diferencia = float(articulo["cantidad_controlada"]) - float(articulo["cantidad_programada"])
            estado = "Completo" if diferencia == 0 else "Pendiente" if diferencia < 0 else "Excedido"
            hoja.append([
                articulo["sku"], articulo["descripcion"], articulo["cantidad_programada"],
                articulo["cantidad_controlada"], diferencia, estado
            ])
        for letra, ancho in zip("ABCDEF", [25, 55, 22, 22, 16, 16]):
            hoja.column_dimensions[letra].width = ancho
        hoja.freeze_panes = "A2"
        archivo = BytesIO()
        libro.save(archivo)
        archivo.seek(0)
        return send_file(
            archivo, as_attachment=True,
            download_name=f"control_embarque_{embarque_id}.xlsx",
            mimetype=("application/vnd.openxmlformats-officedocument."
                      "spreadsheetml.sheet")
        )
    finally:
        conexion.close()


@embarques_bp.route("/api/embarques/<int:embarque_id>/control/registrar", methods=["POST"])
def registrar_control_embarque(embarque_id):
    if not requerido():
        return jsonify({"ok": False, "mensaje": "Sesión no válida"}), 401

    datos = request.get_json(silent=True) or request.form
    sku = str(datos.get("sku", "")).strip()
    try:
        cantidad = float(datos.get("cantidad", 0))
    except (TypeError, ValueError):
        cantidad = 0

    if not sku or cantidad <= 0:
        return jsonify({"ok": False, "mensaje": "Ingresa un SKU y una cantidad válida."}), 400

    conexion = db()
    try:
        control = conexion.execute("""
            SELECT id, estado FROM controles_embarque WHERE embarque_id = ?
        """, (embarque_id,)).fetchone()
        if control is None or control["estado"] != "En control":
            return jsonify({"ok": False, "mensaje": "El control no está activo."}), 409

        articulo = conexion.execute("""
            SELECT id, cantidad_programada, cantidad_controlada
            FROM articulos_embarque WHERE embarque_id = ? AND sku = ?
        """, (embarque_id, sku)).fetchone()
        if articulo is None:
            return jsonify({"ok": False, "mensaje": "El SKU no pertenece a este embarque."}), 404

        ahora = time.time()
        with conexion:
            conexion.execute("""
                INSERT INTO movimientos_control_embarque
                (control_id, articulo_id, cantidad, registrado_en)
                VALUES (?, ?, ?, ?)
            """, (control["id"], articulo["id"], cantidad, ahora))
            conexion.execute("""
                UPDATE articulos_embarque
                SET cantidad_controlada = cantidad_controlada + ?
                WHERE id = ?
            """, (cantidad, articulo["id"]))
            conexion.execute("""
                UPDATE controles_embarque SET ultimo_ping = ? WHERE id = ?
            """, (ahora, control["id"]))

        return jsonify({"ok": True, "mensaje": "Cantidad registrada correctamente."})
    finally:
        conexion.close()


@embarques_bp.route("/embarques/<int:embarque_id>/control/finalizar", methods=["POST"])
def finalizar_control_embarque(embarque_id):
    if not requerido():
        return redirect("/")

    conexion = db()
    try:
        with conexion:
            resultado = conexion.execute("""
                UPDATE controles_embarque
                SET estado = 'Finalizado', finalizado_en = ?, ultimo_ping = ?
                WHERE embarque_id = ? AND estado = 'En control'
            """, (time.time(), time.time(), embarque_id))
        texto_mensaje = (
            "Control de ingreso finalizado." if resultado.rowcount
            else "No hay un control activo para finalizar."
        )
        return redirect(f"/embarques/{embarque_id}/control?mensaje=" + quote(texto_mensaje))
    finally:
        conexion.close()


@embarques_bp.route("/api/embarques/resumen")
def resumen_embarques():
    if not requerido():
        return jsonify({"error": "Sesion no valida"}), 401

    conexion = db()

    try:
        codigo_empresa = empresa_actual()
        en_proceso = conexion.execute("""
            SELECT COUNT(*) AS total
            FROM embarques
            WHERE estado = 'En proceso' AND (? = '' OR empresa_codigo = ?)
        """, (codigo_empresa, codigo_empresa)).fetchone()["total"]

        proximos = conexion.execute("""
            SELECT COUNT(*) AS total
            FROM embarques
            WHERE estado = 'Proximo' AND (? = '' OR empresa_codigo = ?)
        """, (codigo_empresa, codigo_empresa)).fetchone()["total"]

        ingresos = conexion.execute("""
            SELECT fecha, proveedor, estado
            FROM embarques
            WHERE (? = '' OR empresa_codigo = ?)
            ORDER BY fecha ASC
        """, (codigo_empresa, codigo_empresa)).fetchall()

        return jsonify({
            "en_proceso": en_proceso,
            "proximos": proximos,
            "ingresos": [
                {
                    "fecha": ingreso["fecha"],
                    "proveedor": ingreso["proveedor"],
                    "estado": ingreso["estado"]
                }
                for ingreso in ingresos
            ]
        })
    finally:
        conexion.close()
