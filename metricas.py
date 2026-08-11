from flask import Blueprint, render_template, request, redirect, session, jsonify, send_file
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from urllib.parse import quote, urlencode
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta
from io import BytesIO
from database import conectar
import hashlib
import os
import re
import time
import unicodedata


metricas_bp = Blueprint("metricas", __name__)
RUTA_BD = os.environ.get(
    "POLO_OESTE_DB",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "actividad.db")
)

TAREAS = ("Picking", "Almacenaje", "Expedición")
APLICACIONES_ALMACENAJE = {"REC522", "REC510"}
APLICACION_PICKING_TRACE = "PRD095"
APLICACION_EXPEDICION = "EXP040"
APLICACION_RECEPCION = "REC501"
# El WMS contiene varios eventos técnicos por una misma ubicación.  El Trace
# no permite distinguirlos, por lo que se toma una cadencia operativa prudente
# de dos minutos y nunca se asume que una persona estuvo productiva el 100 %
# del tiempo que permaneció activa.
MINUTOS_POR_UBICACION_PICKING = 2
OCUPACION_MAXIMA_ESTIMADA_PICKING = 0.80


def db():
    return conectar(RUTA_BD)


def requerido():
    return "usuario" in session


def redirigir_metricas(mensaje):
    return redirect("/metricas?mensaje=" + quote(mensaje))


def texto(valor):
    return str(valor or "").strip()


def empresa_en_sesion():
    return texto(session.get("empresa_codigo"))


def condicion_empresa(campo="empresa_codigo"):
    codigo = empresa_en_sesion()
    return (f" AND {campo} = ?", [codigo]) if codigo else ("", [])


def sincronizar_operativas_funcionarios(conexion):
    """Acumula las operativas detectadas en Trace para cada cuenta operativa.

    No se eliminan asignaciones anteriores: si el funcionario aparece mañana en
    otra empresa, esta se suma automáticamente a su ficha.
    """
    conexion.execute("""
        INSERT OR IGNORE INTO usuario_empresas
        (usuario_id, empresa_codigo, origen, asignado_en)
        SELECT usuarios.id, movimientos_metricas.empresa_codigo, 'Trace de Stock', ?
        FROM usuarios
        INNER JOIN movimientos_metricas
            ON LOWER(TRIM(movimientos_metricas.funcionario)) =
               LOWER(TRIM(COALESCE(usuarios.nombre_funcionario, usuarios.usuario)))
        WHERE usuarios.es_admin = 0
          AND movimientos_metricas.empresa_codigo IS NOT NULL
          AND movimientos_metricas.empresa_codigo != ''
    """, (time.time(),))


def nombre_normalizado(valor):
    """Clave estable para cruzar el nombre del WMS con Actividad."""
    valor = unicodedata.normalize("NFD", texto(valor).lower())
    valor = "".join(letra for letra in valor if unicodedata.category(letra) != "Mn")
    return re.sub(r"\s+", " ", valor).strip()


def tiempo_legible(segundos):
    segundos = max(0, int(segundos or 0))
    horas, resto = divmod(segundos, 3600)
    minutos = resto // 60
    return f"{horas:02d}:{minutos:02d}"


def limites_del_mes(mes):
    inicio = datetime.strptime(mes, "%Y-%m")
    if inicio.month == 12:
        siguiente = inicio.replace(year=inicio.year + 1, month=1)
    else:
        siguiente = inicio.replace(month=inicio.month + 1)
    return inicio.timestamp(), siguiente.timestamp()


def encabezado(valor):
    valor = unicodedata.normalize("NFD", texto(valor).lower())
    valor = "".join(letra for letra in valor if unicodedata.category(letra) != "Mn")
    return re.sub(r"[^a-z0-9]", "", valor)


def numero(valor):
    if valor is None or valor == "":
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    limpio = texto(valor).replace(".", "").replace(",", ".")
    try:
        return float(limpio)
    except ValueError:
        return 0.0


def fecha_operativa(valor):
    if isinstance(valor, datetime):
        return valor.date().isoformat()

    if isinstance(valor, date):
        return valor.isoformat()

    valor = texto(valor)
    for formato in (
        "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y",
        "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"
    ):
        try:
            return datetime.strptime(valor, formato).date().isoformat()
        except ValueError:
            continue

    return ""


def valor_fila(fila, columnas, nombre):
    indice = columnas.get(nombre)
    return fila[indice] if indice is not None and indice < len(fila) else None


def clave_segura(prefijo, partes):
    texto_base = "|".join(texto(parte) for parte in partes)
    return prefijo + ":" + hashlib.sha256(texto_base.encode("utf-8")).hexdigest()


def crear_tablas():
    conexion = db()
    try:
        conexion.execute("""
            CREATE TABLE IF NOT EXISTS cargas_metricas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tipo TEXT NOT NULL,
                archivo TEXT NOT NULL,
                cargado_en REAL NOT NULL,
                filas_leidas INTEGER NOT NULL DEFAULT 0,
                registros_nuevos INTEGER NOT NULL DEFAULT 0,
                repetidos INTEGER NOT NULL DEFAULT 0,
                omitidos INTEGER NOT NULL DEFAULT 0
            )
        """)
        conexion.execute("""
            CREATE TABLE IF NOT EXISTS movimientos_metricas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                clave_origen TEXT NOT NULL UNIQUE,
                carga_id INTEGER NOT NULL,
                fuente TEXT NOT NULL,
                tarea TEXT NOT NULL,
                fecha TEXT NOT NULL,
                funcionario_codigo TEXT,
                funcionario TEXT NOT NULL,
                unidades REAL NOT NULL,
                empresa_codigo TEXT,
                empresa TEXT,
                aplicacion TEXT,
                referencia TEXT,
                creado_en REAL NOT NULL,
                FOREIGN KEY (carga_id) REFERENCES cargas_metricas(id)
            )
        """)
        conexion.execute("""
            CREATE INDEX IF NOT EXISTS indice_metricas_fecha
            ON movimientos_metricas(fecha, tarea)
        """)
        conexion.execute("""
            CREATE INDEX IF NOT EXISTS indice_metricas_funcionario
            ON movimientos_metricas(funcionario, tarea)
        """)
        columnas_movimientos = {
            columna["name"]
            for columna in conexion.execute("PRAGMA table_info(movimientos_metricas)")
        }
        if "empresa_codigo" not in columnas_movimientos:
            conexion.execute("ALTER TABLE movimientos_metricas ADD COLUMN empresa_codigo TEXT")
        equivalencias = {
            "POLAKOF": "6009", "DECATHLON": "6000", "FORUM": "6005",
            "UNILOG": "6003", "KIABI": "6002", "TELESHOPPING": "6001"
        }
        for descripcion, codigo in equivalencias.items():
            conexion.execute("""
                UPDATE movimientos_metricas SET empresa_codigo = ?
                WHERE (empresa_codigo IS NULL OR empresa_codigo = '')
                  AND UPPER(empresa) = ?
            """, (codigo, descripcion))
        # Desde esta versión el Trace de Stock es la fuente única para Picking.
        # Se conservan las cargas históricas anteriores, pero no se mezclan con
        # el indicador operativo actual para evitar duplicar resultados.
        conexion.execute("""
            UPDATE movimientos_metricas
            SET tarea = 'Picking (histórico)'
            WHERE fuente = 'Consulta de Picking' AND tarea = 'Picking'
        """)
        conexion.execute("""
            UPDATE movimientos_metricas
            SET tarea = 'Picking'
            WHERE tarea = 'Picking Trace' AND aplicacion = ?
        """, (APLICACION_PICKING_TRACE,))
        conexion.commit()
    finally:
        conexion.close()


def leer_trace(archivo):
    libro = load_workbook(archivo, read_only=True, data_only=True)
    hoja = libro.active
    filas = hoja.iter_rows(values_only=True)
    cabecera = next(filas, None)

    if not cabecera:
        raise ValueError("El archivo no contiene encabezados.")

    columnas = {encabezado(nombre): indice for indice, nombre in enumerate(cabecera)}
    requeridas = {"numtrace", "alta", "aplicacion", "codfun", "nombre", "movimiento"}
    faltantes = requeridas - set(columnas)
    if faltantes:
        raise ValueError("El Trace de Stock no tiene las columnas requeridas.")

    movimientos = []
    omitidos = 0
    leidas = 0

    for fila in filas:
        if not any(valor is not None and texto(valor) for valor in fila):
            continue

        leidas += 1
        aplicacion = texto(valor_fila(fila, columnas, "aplicacion")).upper()
        if aplicacion in APLICACIONES_ALMACENAJE:
            tarea = "Almacenaje"
        elif aplicacion == APLICACION_PICKING_TRACE:
            # PRD095 es la evidencia de resultado. Nunca se usa para calcular
            # ni cerrar tiempo: ese dato pertenece exclusivamente a Actividad.
            tarea = "Picking"
        elif aplicacion == APLICACION_EXPEDICION:
            tarea = "Expedición"
        elif aplicacion == APLICACION_RECEPCION:
            tarea = "Recepción"
        else:
            omitidos += 1
            continue

        fecha = fecha_operativa(valor_fila(fila, columnas, "alta"))
        funcionario = texto(valor_fila(fila, columnas, "nombre"))
        unidades = abs(numero(valor_fila(fila, columnas, "movimiento")))
        numero_trace = texto(valor_fila(fila, columnas, "numtrace"))

        if not fecha or not funcionario or not numero_trace or unidades <= 0:
            omitidos += 1
            continue

        movimientos.append({
            "clave": clave_segura("trace", [numero_trace]),
            "fuente": "Trace de Stock",
            "tarea": tarea,
            "fecha": fecha,
            "codigo": texto(valor_fila(fila, columnas, "codfun")),
            "funcionario": funcionario,
            "unidades": unidades,
            "empresa_codigo": texto(valor_fila(fila, columnas, "codemp")),
            "empresa": texto(valor_fila(fila, columnas, "empresa")),
            "aplicacion": aplicacion,
            "referencia": numero_trace
        })

    return movimientos, leidas, omitidos


def leer_picking(archivo):
    libro = load_workbook(archivo, read_only=True, data_only=True)
    hoja = libro.active
    filas = hoja.iter_rows(values_only=True)
    cabecera = next(filas, None)

    if not cabecera:
        raise ValueError("El archivo no contiene encabezados.")

    columnas = {encabezado(nombre): indice for indice, nombre in enumerate(cabecera)}
    requeridas = {
        "preparacion", "producto", "ubicacion", "cantidadprep",
        "pickeo", "funcpick", "descfuncpick"
    }
    faltantes = requeridas - set(columnas)
    if faltantes:
        raise ValueError("La Consulta de Picking no tiene las columnas requeridas.")

    movimientos = []
    omitidos = 0
    leidas = 0

    for fila in filas:
        if not any(valor is not None and texto(valor) for valor in fila):
            continue

        leidas += 1
        fecha = fecha_operativa(valor_fila(fila, columnas, "pickeo"))
        funcionario = texto(valor_fila(fila, columnas, "descfuncpick"))
        unidades = abs(numero(valor_fila(fila, columnas, "cantidadprep")))
        preparacion = texto(valor_fila(fila, columnas, "preparacion"))
        producto = texto(valor_fila(fila, columnas, "producto"))
        ubicacion = texto(valor_fila(fila, columnas, "ubicacion"))
        momento_pickeo = texto(valor_fila(fila, columnas, "pickeo"))
        contenedor = texto(valor_fila(fila, columnas, "nrocontenedor"))

        if not fecha or not funcionario or not preparacion or unidades <= 0:
            omitidos += 1
            continue

        movimientos.append({
            "clave": clave_segura(
                "picking",
                [preparacion, producto, ubicacion, contenedor, momento_pickeo, funcionario, unidades]
            ),
            "fuente": "Consulta de Picking",
            "tarea": "Picking",
            "fecha": fecha,
            "codigo": texto(valor_fila(fila, columnas, "funcpick")),
            "funcionario": funcionario,
            "unidades": unidades,
            "empresa_codigo": texto(valor_fila(fila, columnas, "codemp")),
            "empresa": texto(valor_fila(fila, columnas, "empresa")),
            "aplicacion": "PICKING",
            "referencia": preparacion
        })

    return movimientos, leidas, omitidos


def mes_valido(mes):
    try:
        datetime.strptime(mes, "%Y-%m")
        return True
    except ValueError:
        return False


def filtro_mes(mes, incluir_solo_tareas=False):
    where = "WHERE substr(fecha, 1, 7) = ?"
    parametros = [mes]
    if incluir_solo_tareas:
        where += " AND tarea IN ('Picking', 'Almacenaje', 'Expedición')"
    empresa_sql, empresa_parametros = condicion_empresa()
    return where + empresa_sql, parametros + empresa_parametros


def filtros_detalle(tarea, fecha_desde, fecha_hasta, funcionario):
    condiciones = ["tarea = ?"]
    parametros = [tarea]

    if fecha_desde:
        condiciones.append("fecha >= ?")
        parametros.append(fecha_desde)
    if fecha_hasta:
        condiciones.append("fecha <= ?")
        parametros.append(fecha_hasta)
    if funcionario:
        condiciones.append("funcionario LIKE ? COLLATE NOCASE")
        parametros.append("%" + funcionario + "%")

    codigo = empresa_en_sesion()
    if codigo:
        condiciones.append("empresa_codigo = ?")
        parametros.append(codigo)

    return " AND ".join(condiciones), parametros


@metricas_bp.route("/metricas")
def metricas():
    if not requerido():
        return redirect("/")

    conexion = db()
    try:
        mes = texto(request.args.get("mes"))
        if mes and not mes_valido(mes):
            return redirigir_metricas("El mes elegido no es válido.")

        if not mes:
            empresa_sql, empresa_parametros = condicion_empresa()
            ultimo_dato = conexion.execute(f"""
                SELECT MAX(fecha) AS fecha
                FROM movimientos_metricas
                WHERE 1 = 1 {empresa_sql}
            """, empresa_parametros).fetchone()["fecha"]
            mes = ultimo_dato[:7] if ultimo_dato else date.today().strftime("%Y-%m")

        where, parametros = filtro_mes(mes, incluir_solo_tareas=True)
        resumen = {tarea: 0 for tarea in TAREAS}
        filas_resumen = conexion.execute(f"""
            SELECT tarea, COALESCE(SUM(unidades), 0) AS unidades
            FROM movimientos_metricas {where}
            GROUP BY tarea
        """, parametros).fetchall()
        for fila in filas_resumen:
            if fila["tarea"] in resumen:
                resumen[fila["tarea"]] = int(fila["unidades"])

        empresa_sql, empresa_parametros = condicion_empresa()
        recepcion_mes = conexion.execute(f"""
            SELECT COALESCE(SUM(unidades), 0) AS unidades
            FROM movimientos_metricas
            WHERE tarea = 'Recepción'
              AND substr(fecha, 1, 7) = ?
              {empresa_sql}
        """, [mes] + empresa_parametros).fetchone()["unidades"]

        diarios = conexion.execute(f"""
            SELECT fecha,
                   SUM(CASE WHEN tarea = 'Picking' THEN unidades ELSE 0 END) AS picking,
                   SUM(CASE WHEN tarea = 'Almacenaje' THEN unidades ELSE 0 END) AS almacenaje,
                   SUM(CASE WHEN tarea = 'Expedición' THEN unidades ELSE 0 END) AS expedicion,
                   COUNT(DISTINCT funcionario_codigo || '|' || funcionario) AS funcionarios
            FROM movimientos_metricas {where}
            GROUP BY fecha
            ORDER BY fecha DESC
            LIMIT 60
        """, parametros).fetchall()

        top_por_tarea = {}
        for tarea in TAREAS:
            filtros = [tarea, mes] + empresa_parametros
            where_tarea = "WHERE tarea = ? AND substr(fecha, 1, 7) = ?" + empresa_sql
            top_por_tarea[tarea] = conexion.execute(f"""
                SELECT funcionario, SUM(unidades) AS unidades
                FROM movimientos_metricas
                {where_tarea}
                GROUP BY funcionario_codigo, funcionario
                ORDER BY unidades DESC, funcionario ASC
                LIMIT 5
            """, filtros).fetchall()

        funcionarios_diarios = conexion.execute(f"""
            SELECT fecha, tarea, funcionario, SUM(unidades) AS unidades
            FROM movimientos_metricas {where}
            GROUP BY fecha, tarea, funcionario_codigo, funcionario
            ORDER BY fecha DESC, tarea ASC, unidades DESC, funcionario ASC
            LIMIT 180
        """, parametros).fetchall()

        cargas = conexion.execute("""
            SELECT * FROM cargas_metricas
            ORDER BY cargado_en DESC
            LIMIT 12
        """).fetchall()

        total_registros = conexion.execute(f"""
            SELECT COUNT(*) AS total
            FROM movimientos_metricas
            WHERE substr(fecha, 1, 7) = ?
              {empresa_sql}
        """, [mes] + empresa_parametros).fetchone()["total"]

        return render_template(
            "metricas.html", mes=mes, resumen=resumen, diarios=diarios,
            top_por_tarea=top_por_tarea, funcionarios_diarios=funcionarios_diarios,
            cargas=cargas, total_registros=total_registros,
            recepcion_mes=int(recepcion_mes)
        )
    finally:
        conexion.close()


@metricas_bp.route("/metricas/exportar")
def exportar_metricas():
    if not requerido():
        return redirect("/")

    mes = texto(request.args.get("mes"))
    if not mes or not mes_valido(mes):
        conexion = db()
        try:
            empresa_sql, empresa_parametros = condicion_empresa()
            ultima = conexion.execute(f"""
                SELECT MAX(fecha) AS fecha FROM movimientos_metricas
                WHERE 1 = 1 {empresa_sql}
            """, empresa_parametros).fetchone()["fecha"]
            mes = ultima[:7] if ultima else date.today().strftime("%Y-%m")
        finally:
            conexion.close()

    conexion = db()
    try:
        empresa_sql, empresa_parametros = condicion_empresa()
        diarios = conexion.execute(f"""
            SELECT fecha,
                   SUM(CASE WHEN tarea = 'Picking' THEN unidades ELSE 0 END) AS picking,
                   SUM(CASE WHEN tarea = 'Almacenaje' THEN unidades ELSE 0 END) AS almacenaje,
                   SUM(CASE WHEN tarea = 'Expedición' THEN unidades ELSE 0 END) AS expedicion,
                   COUNT(DISTINCT funcionario_codigo || '|' || funcionario) AS funcionarios
            FROM movimientos_metricas
            WHERE substr(fecha, 1, 7) = ?
              AND tarea IN ('Picking', 'Almacenaje', 'Expedición')
              {empresa_sql}
            GROUP BY fecha ORDER BY fecha ASC
        """, [mes] + empresa_parametros).fetchall()
        funcionarios = conexion.execute(f"""
            SELECT fecha, tarea, funcionario, SUM(unidades) AS unidades
            FROM movimientos_metricas
            WHERE substr(fecha, 1, 7) = ?
              AND tarea IN ('Picking', 'Almacenaje', 'Expedición')
              {empresa_sql}
            GROUP BY fecha, tarea, funcionario_codigo, funcionario
            ORDER BY fecha ASC, tarea ASC, funcionario ASC
        """, [mes] + empresa_parametros).fetchall()

        libro = Workbook()
        hoja = libro.active
        hoja.title = "Resumen diario"
        hoja.append(["Fecha", "Picking", "Almacenaje", "Expedición", "Funcionarios"])
        for fila in diarios:
            hoja.append([fila["fecha"], fila["picking"], fila["almacenaje"], fila["expedicion"], fila["funcionarios"]])
        detalle = libro.create_sheet("Funcionarios")
        detalle.append(["Fecha", "Tarea", "Funcionario", "Unidades"])
        for fila in funcionarios:
            detalle.append([fila["fecha"], fila["tarea"], fila["funcionario"], fila["unidades"]])
        for pagina in (hoja, detalle):
            pagina.freeze_panes = "A2"
            for celda in pagina[1]:
                celda.font = Font(bold=True)
        archivo = BytesIO()
        libro.save(archivo)
        archivo.seek(0)
        return send_file(
            archivo, as_attachment=True, download_name=f"metricas_{mes}.xlsx",
            mimetype=("application/vnd.openxmlformats-officedocument."
                      "spreadsheetml.sheet")
        )
    finally:
        conexion.close()


def rango_actividad_productividad(mes, fecha="", fecha_desde="", fecha_hasta=""):
    """Devuelve los límites Unix para los filtros de Actividad.

    Si no se informan fechas se conserva el comportamiento habitual: mostrar
    el último mes operativo. Cuando se informa una fecha o un período, ese
    filtro trabaja por sí mismo y puede combinarse con los demás.
    """
    if not (fecha or fecha_desde or fecha_hasta):
        return limites_del_mes(mes)

    inicio = 0
    fin = datetime(2100, 1, 1).timestamp()

    def limites_dia(valor):
        try:
            dia = datetime.strptime(valor, "%Y-%m-%d")
        except ValueError:
            return None
        return dia.timestamp(), (dia + timedelta(days=1)).timestamp()

    limite_fecha = limites_dia(fecha) if fecha else None
    limite_desde = limites_dia(fecha_desde) if fecha_desde else None
    limite_hasta = limites_dia(fecha_hasta) if fecha_hasta else None

    if limite_fecha:
        inicio = max(inicio, limite_fecha[0])
        fin = min(fin, limite_fecha[1])
    if limite_desde:
        inicio = max(inicio, limite_desde[0])
    if limite_hasta:
        fin = min(fin, limite_hasta[1])
    return inicio, fin


def actividad_por_funcionario(
    conexion, mes, fecha="", fecha_desde="", fecha_hasta="", funcionario=""
):
    """Suma períodos de Actividad que se superponen con el mes indicado.

    El Trace no interviene aquí: puede haber un período activo sin movimientos
    WMS y debe conservarse tal como fue informado por el handheld.
    """
    inicio_mes, fin_mes = rango_actividad_productividad(
        mes, fecha, fecha_desde, fecha_hasta
    )
    ahora = time.time()
    condiciones_funcionario = ""
    parametros_funcionario = []
    if funcionario:
        condiciones_funcionario = " AND actividades.funcionario LIKE ? COLLATE NOCASE"
        parametros_funcionario.append("%" + funcionario + "%")
    periodos = conexion.execute("""
        SELECT actividades.funcionario, periodos_actividad.inicio,
               periodos_actividad.fin
        FROM periodos_actividad
        INNER JOIN actividades
            ON actividades.id = periodos_actividad.actividad_id
        WHERE periodos_actividad.inicio < ?
          AND COALESCE(periodos_actividad.fin, ?) > ?
    """ + condiciones_funcionario, [fin_mes, ahora, inicio_mes] + parametros_funcionario).fetchall()

    resultado = {}
    for periodo in periodos:
        inicio = max(float(periodo["inicio"]), inicio_mes)
        fin = min(float(periodo["fin"] or ahora), fin_mes)
        if fin <= inicio:
            continue
        clave = nombre_normalizado(periodo["funcionario"])
        cursor = inicio
        while cursor < fin:
            siguiente_dia = datetime.fromtimestamp(cursor).replace(
                hour=0, minute=0, second=0, microsecond=0
            ).timestamp() + 86400
            corte = min(fin, siguiente_dia)
            fecha = datetime.fromtimestamp(cursor).date().isoformat()
            detalle = resultado.setdefault(clave, {
                "segundos": 0, "dias": {}, "funcionario": periodo["funcionario"]
            })
            dia = detalle["dias"].setdefault(fecha, {"segundos": 0, "periodos": 0})
            dia["segundos"] += corte - cursor
            dia["periodos"] += 1
            detalle["segundos"] += corte - cursor
            cursor = corte
    return resultado


def resumen_productividad_picking(
    conexion, mes, fecha="", fecha_desde="", fecha_hasta="", funcionario=""
):
    empresa_sql, empresa_parametros = condicion_empresa()
    condiciones = ["tarea = 'Picking'", "aplicacion = ?"]
    parametros = [APLICACION_PICKING_TRACE]
    if fecha:
        condiciones.append("fecha = ?")
        parametros.append(fecha)
    if fecha_desde:
        condiciones.append("fecha >= ?")
        parametros.append(fecha_desde)
    if fecha_hasta:
        condiciones.append("fecha <= ?")
        parametros.append(fecha_hasta)
    if not (fecha or fecha_desde or fecha_hasta):
        condiciones.append("substr(fecha, 1, 7) = ?")
        parametros.append(mes)
    if funcionario:
        condiciones.append("funcionario LIKE ? COLLATE NOCASE")
        parametros.append("%" + funcionario + "%")

    where_trace = " AND ".join(condiciones) + empresa_sql
    filas_trace = conexion.execute(f"""
        SELECT funcionario_codigo, funcionario,
               COALESCE(SUM(unidades), 0) AS unidades,
               COUNT(*) AS registros,
               GROUP_CONCAT(DISTINCT fecha) AS fechas_trace
        FROM movimientos_metricas
        WHERE {where_trace}
        GROUP BY funcionario_codigo, funcionario
        ORDER BY unidades DESC, funcionario ASC
    """, parametros + empresa_parametros).fetchall()

    tiempos = actividad_por_funcionario(
        conexion, mes, fecha, fecha_desde, fecha_hasta, funcionario
    )
    filas = []
    funcionarios_trace = set()
    for fila in filas_trace:
        clave_funcionario = nombre_normalizado(fila["funcionario"])
        funcionarios_trace.add(clave_funcionario)
        detalle_tiempo = tiempos.get(clave_funcionario, {"segundos": 0, "dias": {}})
        segundos = int(detalle_tiempo["segundos"])
        dias = detalle_tiempo["dias"]
        fechas_trace = set(filter(None, texto(fila["fechas_trace"]).split(",")))
        fechas_jornada = fechas_trace | set(dias)
        pausas = sum(max(0, int(dia["periodos"]) - 1) for dia in dias.values())
        efectivo_8h = int(sum(min(dia["segundos"], 8 * 3600) for dia in dias.values()))
        capacidad_8h = len(dias) * 8 * 3600
        unidades = int(fila["unidades"] or 0)
        registros = int(fila["registros"] or 0)
        # El Trace no conserva una hora útil para cada ubicación y puede
        # registrar varios eventos técnicos para un mismo movimiento. Se usa
        # una referencia de 2 minutos por ubicación, limitada al 80 % de la
        # actividad real: así el indicador no convierte el ruido del Trace en
        # una jornada ficticia de productividad perfecta.
        capacidad_realista = int(segundos * OCUPACION_MAXIMA_ESTIMADA_PICKING)
        productivo_estimado = min(
            capacidad_realista,
            registros * MINUTOS_POR_UBICACION_PICKING * 60
        )
        sin_produccion = max(0, segundos - productivo_estimado)
        filas.append({
            "funcionario": fila["funcionario"],
            "codigo": fila["funcionario_codigo"] or "-",
            "unidades": unidades,
            "registros": registros,
            "segundos": segundos,
            "tiempo": tiempo_legible(segundos),
            "unidades_hora": round(unidades / (segundos / 3600), 1) if segundos else None,
            "con_actividad": segundos > 0,
            "sin_produccion_segundos": sin_produccion,
            "sin_produccion": tiempo_legible(sin_produccion),
            "productivo_estimado": tiempo_legible(productivo_estimado),
            "pausas": pausas,
            "efectivo_8h_segundos": efectivo_8h,
            "efectivo_8h": tiempo_legible(efectivo_8h),
            "efectividad_8h": round((efectivo_8h / capacidad_8h) * 100, 1) if capacidad_8h else None,
            "dias_activos": len(dias),
            "jornadas": len(fechas_jornada),
            "fechas_jornada": sorted(fechas_jornada)
        })

    # Mantiene visibles las jornadas activas sin resultado PRD095, que son
    # justamente las que requieren revisión del supervisor.
    for clave_funcionario, detalle_tiempo in tiempos.items():
        if clave_funcionario in funcionarios_trace:
            continue
        segundos = int(detalle_tiempo["segundos"])
        dias = detalle_tiempo["dias"]
        efectivo_8h = int(sum(min(dia["segundos"], 8 * 3600) for dia in dias.values()))
        capacidad_8h = len(dias) * 8 * 3600
        pausas = sum(max(0, int(dia["periodos"]) - 1) for dia in dias.values())
        filas.append({
            "funcionario": detalle_tiempo.get("funcionario", "Sin identificar"),
            "codigo": "-", "unidades": 0, "registros": 0,
            "segundos": segundos, "tiempo": tiempo_legible(segundos),
            "unidades_hora": None, "con_actividad": segundos > 0,
            "sin_produccion_segundos": segundos,
            "sin_produccion": tiempo_legible(segundos),
            "productivo_estimado": tiempo_legible(0),
            "pausas": pausas, "efectivo_8h_segundos": efectivo_8h,
            "efectivo_8h": tiempo_legible(efectivo_8h),
            "efectividad_8h": round((efectivo_8h / capacidad_8h) * 100, 1) if capacidad_8h else None,
            "dias_activos": len(dias),
            "jornadas": len(dias),
            "fechas_jornada": sorted(dias)
        })

    filas.sort(key=lambda item: (-item["unidades"], item["funcionario"].lower()))

    return filas


@metricas_bp.route("/productividad-picking")
def productividad_picking():
    if not requerido():
        return redirect("/")

    conexion = db()
    try:
        mes = texto(request.args.get("mes"))
        if mes and not mes_valido(mes):
            return redirect("/productividad-picking")

        if not mes:
            empresa_sql, empresa_parametros = condicion_empresa()
            ultima_fecha = conexion.execute(f"""
                SELECT MAX(fecha) AS fecha
                FROM movimientos_metricas
                WHERE tarea = 'Picking' AND aplicacion = ? {empresa_sql}
            """, [APLICACION_PICKING_TRACE] + empresa_parametros).fetchone()["fecha"]
            mes = ultima_fecha[:7] if ultima_fecha else date.today().strftime("%Y-%m")

        fecha = texto(request.args.get("fecha"))
        fecha_desde = texto(request.args.get("fecha_desde"))
        fecha_hasta = texto(request.args.get("fecha_hasta"))
        funcionario = texto(request.args.get("funcionario"))

        filas = resumen_productividad_picking(
            conexion, mes, fecha, fecha_desde, fecha_hasta, funcionario
        )
        total_unidades = sum(fila["unidades"] for fila in filas)
        total_registros = sum(fila["registros"] for fila in filas)
        con_actividad = sum(1 for fila in filas if fila["con_actividad"])
        total_segundos = sum(fila["segundos"] for fila in filas)
        sin_produccion_total = sum(fila["sin_produccion_segundos"] for fila in filas)
        pausas_totales = sum(fila["pausas"] for fila in filas)
        efectivo_8h_total = sum(fila["efectivo_8h_segundos"] for fila in filas)
        jornadas_analizadas = len({
            fecha_jornada
            for fila in filas
            for fecha_jornada in fila["fechas_jornada"]
        })
        parametros_filtro = {
            clave: valor for clave, valor in {
                "mes": mes,
                "fecha": fecha,
                "fecha_desde": fecha_desde,
                "fecha_hasta": fecha_hasta,
                "funcionario": funcionario
            }.items() if valor
        }
        url_exportacion = "/productividad-picking/exportar?" + urlencode(
            parametros_filtro
        )
        url_vuelta = "/productividad-picking?" + urlencode(parametros_filtro)

        if fecha:
            periodo_etiqueta = "Jornada seleccionada: " + fecha
        elif fecha_desde or fecha_hasta:
            periodo_etiqueta = "Periodo seleccionado"
        else:
            periodo_etiqueta = "Ultimo mes con datos: " + mes
        return render_template(
            "productividad_picking.html", mes=mes, filas=filas,
            fecha=fecha, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            funcionario=funcionario, periodo_etiqueta=periodo_etiqueta,
            jornadas_analizadas=jornadas_analizadas,
            url_exportacion=url_exportacion, url_vuelta=url_vuelta,
            total_unidades=total_unidades, total_registros=total_registros,
            con_actividad=con_actividad, total_segundos=total_segundos,
            tiempo_total=tiempo_legible(total_segundos),
            sin_produccion_total=tiempo_legible(sin_produccion_total),
            pausas_totales=pausas_totales,
            efectivo_8h_total=tiempo_legible(efectivo_8h_total)
        )
    finally:
        conexion.close()


@metricas_bp.route("/productividad-picking/exportar")
def exportar_productividad_picking():
    if not requerido():
        return redirect("/")

    conexion = db()
    try:
        mes = texto(request.args.get("mes"))
        if mes and not mes_valido(mes):
            return redirect("/productividad-picking")
        if not mes:
            empresa_sql, empresa_parametros = condicion_empresa()
            ultima_fecha = conexion.execute(f"""
                SELECT MAX(fecha) AS fecha
                FROM movimientos_metricas
                WHERE tarea = 'Picking' AND aplicacion = ? {empresa_sql}
            """, [APLICACION_PICKING_TRACE] + empresa_parametros).fetchone()["fecha"]
            mes = ultima_fecha[:7] if ultima_fecha else date.today().strftime("%Y-%m")

        fecha = texto(request.args.get("fecha"))
        fecha_desde = texto(request.args.get("fecha_desde"))
        fecha_hasta = texto(request.args.get("fecha_hasta"))
        funcionario = texto(request.args.get("funcionario"))
        filas = resumen_productividad_picking(
            conexion, mes, fecha, fecha_desde, fecha_hasta, funcionario
        )
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Productividad Picking"
        hoja.append([
            "Funcionario", "Código", "Tiempo de actividad", "Unidades PRD095",
            "Registros", "Productividad hora/hombre", "Tiempo productivo estimado",
            "Tiempo activo sin producción",
            "Pausas", "Tiempo efectivo (máx. 8h/día)", "% de jornada efectiva",
            "Jornadas consideradas", "Estado de cruce"
        ])
        for fila in filas:
            hoja.append([
                fila["funcionario"], fila["codigo"], fila["tiempo"], fila["unidades"],
                fila["registros"], fila["unidades_hora"] or "", fila["productivo_estimado"], fila["sin_produccion"],
                fila["pausas"], fila["efectivo_8h"], fila["efectividad_8h"] or "",
                fila["jornadas"],
                "Con actividad" if fila["con_actividad"] else "Sin actividad"
            ])
        for letra, ancho in zip("ABCDEFGHIJKLMN", [28, 16, 20, 18, 14, 23, 22, 24, 12, 25, 20, 18, 18, 18]):
            hoja.column_dimensions[letra].width = ancho
        hoja.freeze_panes = "A2"
        for celda in hoja[1]:
            celda.font = Font(bold=True)
        archivo = BytesIO()
        libro.save(archivo)
        archivo.seek(0)
        return send_file(
            archivo, as_attachment=True,
            download_name=f"productividad_picking_{mes}.xlsx",
            mimetype=("application/vnd.openxmlformats-officedocument."
                      "spreadsheetml.sheet")
        )
    finally:
        conexion.close()


@metricas_bp.route("/metricas/detalle/<tarea>")
def detalle_metrica(tarea):
    if not requerido():
        return redirect("/")

    tareas_disponibles = {
        "picking": "Picking",
        "almacenaje": "Almacenaje"
    }
    tarea_visible = tareas_disponibles.get(tarea)
    if tarea_visible is None:
        return redirect("/metricas")

    fecha_desde = texto(request.args.get("fecha_desde"))
    fecha_hasta = texto(request.args.get("fecha_hasta"))
    funcionario = texto(request.args.get("funcionario"))
    consultado = bool(fecha_desde or fecha_hasta or funcionario)
    registros = []
    total_unidades = 0
    total_coincidencias = 0
    resumen_funcionarios = []

    if consultado:
        conexion = db()
        try:
            clausula, parametros = filtros_detalle(
                tarea_visible, fecha_desde, fecha_hasta, funcionario
            )
            totales = conexion.execute(f"""
                SELECT COUNT(*) AS cantidad, COALESCE(SUM(unidades), 0) AS unidades
                FROM movimientos_metricas
                WHERE {clausula}
            """, parametros).fetchone()
            total_coincidencias = int(totales["cantidad"])
            total_unidades = int(totales["unidades"])
            resumen_funcionarios = conexion.execute(f"""
                SELECT funcionario, funcionario_codigo,
                       SUM(unidades) AS unidades,
                       COUNT(*) AS cantidad_registros
                FROM movimientos_metricas
                WHERE {clausula}
                GROUP BY funcionario_codigo, funcionario
                ORDER BY unidades DESC, cantidad_registros DESC, funcionario ASC
            """, parametros).fetchall()
            registros = conexion.execute(f"""
                SELECT fecha, funcionario, unidades, fuente, aplicacion, referencia
                FROM movimientos_metricas
                WHERE {clausula}
                ORDER BY fecha DESC, funcionario ASC, id DESC
                LIMIT 5000
            """, parametros).fetchall()
        finally:
            conexion.close()

    parametros_exportacion = {
        clave: valor for clave, valor in {
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,
            "funcionario": funcionario
        }.items() if valor
    }
    url_exportacion = (
        f"/metricas/detalle/{tarea}/exportar?" + urlencode(parametros_exportacion)
        if consultado else ""
    )

    return render_template(
        "detalle_metrica.html", tarea=tarea_visible, clave_tarea=tarea,
        fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
        funcionario=funcionario, consultado=consultado,
        registros=registros, total_unidades=total_unidades,
        total_coincidencias=total_coincidencias,
        resumen_funcionarios=resumen_funcionarios,
        url_exportacion=url_exportacion
    )


@metricas_bp.route("/metricas/detalle/<tarea>/exportar")
def exportar_detalle_metrica(tarea):
    if not requerido():
        return redirect("/")

    tareas_disponibles = {
        "picking": "Picking",
        "almacenaje": "Almacenaje"
    }
    tarea_visible = tareas_disponibles.get(tarea)
    if tarea_visible is None:
        return redirect("/metricas")

    fecha_desde = texto(request.args.get("fecha_desde"))
    fecha_hasta = texto(request.args.get("fecha_hasta"))
    funcionario = texto(request.args.get("funcionario"))

    if not (fecha_desde or fecha_hasta or funcionario):
        return redirect(f"/metricas/detalle/{tarea}")

    clausula, parametros = filtros_detalle(
        tarea_visible, fecha_desde, fecha_hasta, funcionario
    )
    conexion = db()

    try:
        resumen_funcionarios = conexion.execute(f"""
            SELECT funcionario, funcionario_codigo,
                   SUM(unidades) AS unidades,
                   COUNT(*) AS cantidad_registros
            FROM movimientos_metricas
            WHERE {clausula}
            GROUP BY funcionario_codigo, funcionario
            ORDER BY unidades DESC, cantidad_registros DESC, funcionario ASC
        """, parametros).fetchall()

        registros = conexion.execute(f"""
            SELECT tarea, fecha, funcionario_codigo, funcionario, unidades,
                   referencia, aplicacion, empresa, fuente
            FROM movimientos_metricas
            WHERE {clausula}
            ORDER BY fecha DESC, funcionario ASC, id DESC
        """, parametros)

        libro = Workbook(write_only=True)
        hoja = libro.create_sheet("Resumen por funcionario")
        hoja.append([
            "Funcionario", "Código de funcionario", "Unidades", "Cantidad de registros"
        ])

        for fila in resumen_funcionarios:
            hoja.append([
                fila["funcionario"], fila["funcionario_codigo"],
                fila["unidades"], fila["cantidad_registros"]
            ])

        detalle = libro.create_sheet("Detalle completo")
        detalle.append([
            "Tarea", "Fecha", "Código de funcionario", "Funcionario",
            "Unidades", "Referencia", "Aplicación", "Empresa", "Fuente"
        ])

        for registro in registros:
            detalle.append([
                registro["tarea"], registro["fecha"], registro["funcionario_codigo"],
                registro["funcionario"], registro["unidades"], registro["referencia"],
                registro["aplicacion"], registro["empresa"], registro["fuente"]
            ])

        archivo = BytesIO()
        libro.save(archivo)
        archivo.seek(0)
        nombre = f"polo_oeste_{tarea}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            archivo, as_attachment=True, download_name=nombre,
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )
    finally:
        conexion.close()


@metricas_bp.route("/metricas/cargar", methods=["POST"])
def cargar_metricas():
    if not requerido():
        return redirect("/")

    tipo = texto(request.form.get("tipo"))
    archivo = request.files.get("archivo")
    tipos = {
        "trace": ("Trace de Stock", leer_trace)
    }

    if tipo not in tipos:
        return redirigir_metricas("Selecciona el tipo de archivo que vas a cargar.")

    if archivo is None or not archivo.filename.lower().endswith(".xlsx"):
        return redirigir_metricas("Selecciona un archivo Excel con extensión .xlsx.")

    nombre_tipo, lector = tipos[tipo]
    try:
        movimientos, filas_leidas, omitidos = lector(archivo)
    except Exception as error:
        return redirigir_metricas("No fue posible leer el archivo: " + texto(error))

    conexion = db()
    try:
        with conexion:
            cursor = conexion.execute("""
                INSERT INTO cargas_metricas (tipo, archivo, cargado_en, filas_leidas, omitidos)
                VALUES (?, ?, ?, ?, ?)
            """, (
                nombre_tipo, secure_filename(archivo.filename) or nombre_tipo,
                time.time(), filas_leidas, omitidos
            ))
            carga_id = cursor.lastrowid
            nuevos = 0
            repetidos = 0

            for movimiento in movimientos:
                resultado = conexion.execute("""
                    INSERT OR IGNORE INTO movimientos_metricas (
                        clave_origen, carga_id, fuente, tarea, fecha,
                        funcionario_codigo, funcionario, unidades, empresa_codigo, empresa,
                        aplicacion, referencia, creado_en
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    movimiento["clave"], carga_id, movimiento["fuente"],
                    movimiento["tarea"], movimiento["fecha"], movimiento["codigo"],
                    movimiento["funcionario"], movimiento["unidades"], movimiento["empresa_codigo"], movimiento["empresa"],
                    movimiento["aplicacion"], movimiento["referencia"], time.time()
                ))
                if resultado.rowcount:
                    nuevos += 1
                else:
                    repetidos += 1

            conexion.execute("""
                UPDATE cargas_metricas
                SET registros_nuevos = ?, repetidos = ?
                WHERE id = ?
            """, (nuevos, repetidos, carga_id))

            sincronizar_operativas_funcionarios(conexion)

        return redirigir_metricas(
            f"{nombre_tipo} cargado: {nuevos} registros nuevos, {repetidos} repetidos y {omitidos} omitidos."
        )
    finally:
        conexion.close()


@metricas_bp.route("/metricas/carga/<int:carga_id>/eliminar", methods=["POST"])
def eliminar_carga(carga_id):
    if not requerido():
        return redirect("/")

    conexion = db()
    try:
        with conexion:
            conexion.execute("DELETE FROM movimientos_metricas WHERE carga_id = ?", (carga_id,))
            resultado = conexion.execute("DELETE FROM cargas_metricas WHERE id = ?", (carga_id,))
        if resultado.rowcount == 0:
            return redirigir_metricas("No se encontró la carga seleccionada.")
        return redirigir_metricas("Carga eliminada y métricas recalculadas.")
    finally:
        conexion.close()


@metricas_bp.route("/api/metricas/resumen-anterior")
def api_resumen_anterior():
    if not requerido():
        return jsonify({"error": "Sesion no valida"}), 401

    conexion = db()
    try:
        datos = {}
        ultima_fecha_general = None
        empresa_sql, empresa_parametros = condicion_empresa()

        for tarea in TAREAS:
            fecha = conexion.execute(f"""
                SELECT MAX(fecha) AS fecha
                FROM movimientos_metricas
                WHERE tarea = ? {empresa_sql}
            """, [tarea] + empresa_parametros).fetchone()["fecha"]

            unidades = 0
            if fecha:
                unidades = conexion.execute(f"""
                    SELECT COALESCE(SUM(unidades), 0) AS unidades
                    FROM movimientos_metricas
                    WHERE tarea = ? AND fecha = ? {empresa_sql}
                """, [tarea, fecha] + empresa_parametros).fetchone()["unidades"]
                if ultima_fecha_general is None or fecha > ultima_fecha_general:
                    ultima_fecha_general = fecha

            datos[tarea] = {"fecha": fecha, "unidades": int(unidades)}

        fecha_distribucion = conexion.execute(f"""
            SELECT MAX(fecha) AS fecha
            FROM movimientos_metricas
            WHERE tarea IN ('Picking', 'Almacenaje', 'Expedición')
            {empresa_sql}
        """, empresa_parametros).fetchone()["fecha"]
        distribucion = {tarea: 0 for tarea in TAREAS}
        if fecha_distribucion:
            filas_distribucion = conexion.execute(f"""
                SELECT tarea, COALESCE(SUM(unidades), 0) AS unidades
                FROM movimientos_metricas
                WHERE fecha = ?
                  AND tarea IN ('Picking', 'Almacenaje', 'Expedición')
                  {empresa_sql}
                GROUP BY tarea
            """, [fecha_distribucion] + empresa_parametros).fetchall()
            for fila in filas_distribucion:
                distribucion[fila["tarea"]] = int(fila["unidades"])

        return jsonify({
            "fecha_general": ultima_fecha_general,
            "fecha_distribucion": fecha_distribucion,
            "distribucion": distribucion,
            "picking": datos["Picking"],
            "almacenaje": datos["Almacenaje"],
            "expedicion": datos["Expedición"]
        })
    finally:
        conexion.close()

